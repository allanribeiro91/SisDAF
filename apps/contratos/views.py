from django.shortcuts import render, redirect
from django.db.models import Sum, F, Case, When, Max, Value
from django.db import models
from django.http import QueryDict
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage
from django.contrib import auth, messages
from apps.main.models import CustomLog
from apps.usuarios.models import Usuario
from apps.produtos.models import DenominacoesGenericas, ProdutosFarmaceuticos
from apps.fornecedores.models import Fornecedores
from apps.contratos.models import (ContratosArps, ContratosArpsItens, Contratos, 
                                   ContratosObjetos, ContratosParcelas, ContratosEntregas,
                                   ContratosFiscais, Empenhos, EmpenhosItens)
from apps.contratos.forms import (ContratosArpsForm, ContratosArpsItensForm, ContratosForm, 
                                  ContratosObjetosForm, ContratosParcelasForm, ContratosEntregasForm,
                                  ContratosFiscaisForm, EmpenhoForm, EmpenhosItensForm)
from setup.choices import UNIDADE_DAF, MODALIDADE_AQUISICAO, STATUS_ARP, YES_NO, TIPO_COTA
from django.http import JsonResponse, HttpResponse
from datetime import datetime
from django.utils import timezone
import pytz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import json

#timezone
tz = pytz.timezone("America/Sao_Paulo")


#CONTRATOS
def contratos(request):

    lista_modalidades = [('', '')] + MODALIDADE_AQUISICAO
    lista_unidadesdaf = [item for item in UNIDADE_DAF if item[0] not in ['cofisc', 'gabinete']]
    
    tabContratos = Contratos.objects.filter(del_status=False).order_by('-data_publicacao')
    
    conteudo = {
        'lista_unidadesdaf': lista_unidadesdaf,
        'lista_modalidades': lista_modalidades,
        'tabContratos': tabContratos,
    }
    return render(request, 'contratos/contratos.html', conteudo)

def contrato_ficha(request, id_contrato=None):
    if id_contrato:
        contrato = Contratos.objects.get(id=id_contrato)
    else:
        contrato = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if contrato:
            contrato_form = ContratosForm(request.POST, instance=contrato)
            novo_contrato = False
        else:
            contrato_form = ContratosForm(request.POST)
            novo_contrato = True
        
        #Verificar se houve alteração no formulário
        if not contrato_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Passar o objeto Denominação Genérica
        denominacao_id = request.POST.get('denominacao')
        denominacao_instance = DenominacoesGenericas.objects.get(id=denominacao_id)
    
        #Passar o objeto Fornecedor
        fornecedor_id = request.POST.get('fornecedor')
        fornecedor_instance =  Fornecedores.objects.get(id=fornecedor_id)

        #Passar a ARP
        modalidade = request.POST.get('modalidade_aquisicao')
        if modalidade == 'pregao_comarp':
            arp_id = request.POST.get('arp')
            arp_instance = ContratosArps.objects.get(id=arp_id)
        else:
            arp_instance = None

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #lei de licitacao
        lei_licitacao = request.POST.get('ct_lei_licitacao_valor')

        #Atualizar os valores no mutable_post
        modificacoes_post['denominacao'] = denominacao_instance
        modificacoes_post['fornecedor'] = fornecedor_instance
        modificacoes_post['arp'] = arp_instance
        modificacoes_post['lei_licitacao'] = lei_licitacao

        #Criar o formulário com os dados atualizados
        contrato_form = ContratosForm(modificacoes_post, instance=contrato_form.instance)
        
        #salvar
        if contrato_form.is_valid():
            #Salvar o produto
            contrato = contrato_form.save(commit=False)
            contrato.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_id = contrato.id
            log_registro_usuario = contrato.usuario_registro.dp_nome_completo
            log_registro_data = contrato.registro_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_atualizacao_usuario = contrato.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = contrato.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = contrato.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Contratos",
                model='Contratos',
                model_id=contrato.id,
                item_id=0,
                item_descricao="Salvar edição de Contrato.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Contrato (ID {contrato.id}, Número do Contrato: {contrato.numero_contrato}, Denominação: {contrato.denominacao.denominacao}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_id': log_id,
                    'log_registro_usuario': log_registro_usuario,
                    'log_registro_data': log_registro_data,
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': novo_contrato,
                    'redirect_url': reverse('contrato_ficha', args=[contrato.id]),
                })
        else:
            print("Erro formulário Contrato")
            print(contrato_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

    list_objetos = []
    list_parcelas = []

    tab_objetos = None
    tab_parcelas = None
    tab_entregas = None
    tab_fiscais = None
    if contrato:
        form = ContratosForm(instance=contrato)
        tab_objetos = ContratosObjetos.objects.filter(del_status=False, contrato_id=id_contrato)
        tab_parcelas = ContratosParcelas.objects.filter(del_status=False, contrato=id_contrato).order_by('objeto__numero_item', 'numero_parcela')
        tab_entregas = ContratosEntregas.objects.filter(del_status=False, contrato=id_contrato).order_by('parcela__objeto__numero_item', 'parcela__numero_parcela', 'numero_entrega')
        tab_fiscais = ContratosFiscais.objects.filter(del_status=False, contrato=id_contrato).order_by('-data_inicio')
        
        if tab_objetos.count()>0:
            for objeto in tab_objetos:
                objeto_id = objeto.id
                item = objeto.numero_item
                produto = objeto.produto.produto
                item_produto = f'[Item {item}] {produto}'
                list_objetos.append((objeto_id, item_produto))
        
        if tab_parcelas.count()>0:
            for parcela in tab_parcelas:
                parcela_id = parcela.id
                item = parcela.objeto.numero_item
                numero_parcela = parcela.numero_parcela
                produto = parcela.objeto.produto.produto
                item_parcela_produto = f'[Item {item}] [Parcela {numero_parcela}] {produto}'
                list_parcelas.append((parcela_id, item_parcela_produto))
    else:
        form = ContratosForm()
        tab_objetos = None
        tab_parcelas = None

    #Formulários
    form_ct_objeto = ContratosObjetosForm()
    form_ct_parcela = ContratosParcelasForm()
    form_ct_entrega = ContratosEntregasForm()
    form_ct_fiscal = ContratosFiscaisForm()

    return render(request, 'contratos/contrato_ficha.html', {
        'form': form,
        'contrato': contrato,
        'form_objeto': form_ct_objeto,
        'form_parcela': form_ct_parcela,
        'form_ct_entrega': form_ct_entrega,
        'form_fiscal': form_ct_fiscal,
        'tab_objetos': tab_objetos,
        'tab_parcelas': tab_parcelas,
        'tab_entregas': tab_entregas,
        'tab_fiscais': tab_fiscais,
        'list_objetos': list_objetos,
        'list_parcelas': list_parcelas
    })

def contrato_delete(request, id_contrato=None):   
    try:
        contrato = Contratos.objects.get(id=id_contrato)
        contrato.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Contratos",
            model='Contratos',
            model_id=contrato.id,
            item_id=0,
            item_descricao="Deleção de Contrato.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Contrato (ID {contrato.id}, Número: {contrato.numero_contrato}, Denominação: {contrato.denominacao.denominacao}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Contrato deletado com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Contrato não encontrado."
            })

def buscar_arps(request, unidade_daf=None):
    arps = ContratosArps.objects.filter(del_status=False, unidade_daf=unidade_daf, status='vigente').order_by('numero_arp')
    arps_list = list(arps.values('id', 'numero_arp', 'denominacao', 'fornecedor'))
    return JsonResponse({'arps': arps_list})

def buscar_arps_itens(request, id_arp=None):
    arps_itens = ContratosArpsItens.objects.filter(del_status=False, arp__status='vigente', arp__id=id_arp).order_by('numero_item')
    arps_itens_list = []

    for item in arps_itens:
        arps_item_data = {
            'id': item.id,
            'numero_item': item.numero_item,
            'tipo_cota': item.tipo_cota,
            'produto': item.produto.produto,
            'qtd_registrada': item.qtd_registrada,
            'qtd_saldo': item.qtd_saldo()
        }
        arps_itens_list.append(arps_item_data)

    return JsonResponse({'arps_itens': arps_itens_list})

def buscar_contrato(request, id_contrato=None):
    contrato = Contratos.objects.get(id=id_contrato)
    ct_denominacao_id = contrato.denominacao.id
    ct_denominacao = str(contrato.denominacao)
    ct_fornecedor_id = contrato.fornecedor.id
    ct_fornecedor = str(contrato.fornecedor)
    contrato_dados = {
            'denominacao_id': ct_denominacao_id,
            'denominacao_texto': ct_denominacao,
            'fornecedor_id': ct_fornecedor_id,
            'fornecedor_texto': ct_fornecedor,
        }
    return JsonResponse({'contrato': contrato_dados})   

def vincular_itens_arp(request, id_arp=None, id_contrato=None):
    arps_itens = ContratosArpsItens.objects.filter(del_status=False, arp__status='vigente', arp__id=id_arp)
    contrato = Contratos.objects.get(id=id_contrato)

    #salvar os itens da ARP
    for item in arps_itens:
        objeto = ContratosObjetos(
            usuario_registro=request.user.usuario_relacionado,
            registro_data=timezone.now(),
            ult_atual_data=timezone.now(),
            numero_item=item.numero_item,
            fator_embalagem=0,
            valor_unitario=item.valor_unitario(),
            produto=item.produto,
            contrato=contrato,
            arp_item=item,
            observacoes_gerais='Sem observações.',
            del_status=False,
        )
        objeto.save(current_user=request.user.usuario_relacionado)

    #buscar os objetos do contrato
    objetos = ContratosObjetos.objects.filter(id=id_contrato)
    objetos_list = []

    for item in objetos:
        objeto_data = {
            'id': item.id,
            'numero_item': item.numero_item,
            'produto': item.produto.produto,
            'fator_embalagem': item.fator_embalagem,
            'valor_unitario': item.valor_unitario,
            'valor_total': item.valor_total(),
        }
        objetos_list.append(objeto_data)

    return JsonResponse({'objetos': objetos_list})

def contrato_dados_arp(request, id_arp=None):
    arps_itens = ContratosArpsItens.objects.filter(del_status=False, arp_id=id_arp).order_by('numero_item')
    if arps_itens:
        arp_numero = arps_itens.first().arp.numero_arp
    else:
        arp_numero = None
    conteudo = {
        'arps_itens': arps_itens,
        'arp_numero': arp_numero,
    }
    return render(request, 'contratos/contrato_ficha_arp.html', conteudo)




#CONTRATOS/OBJETOS
def contrato_objeto_modal(request, id_objeto=None):
    try:
        item = ContratosObjetos.objects.get(id=id_objeto)
        produto_id = item.produto_id
        arp_item_id = item.arp_item_id
        print("ARP Item: ", arp_item_id)
        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'numero_item': item.numero_item,
            'produto': produto_id,
            'arp_item': arp_item_id,
            'fator_embalagem': item.fator_embalagem,
            'parcelas': item.numero_parcelas(),

            #parcelas
            'numero_parcelas': item.numero_parcelas(),
            'parcelas_entregues': item.numero_parcelas_entregues(),
            'parcelas_atraso': item.numero_parcelas_atraso(),
            'parcelas_dias_atraso': item.dias_atraso(),

            #quantidades
            'qtd_contratada': item.qtd_contratada(),
            'qtd_doada': item.qtd_doada_objeto(),
            'qtd_total': item.qtd_total_objeto(),
            'qtd_entregue': item.qtd_entregue(),
            'qtd_a_entregar': item.qtd_a_entregar(),

            #empenho
            'qtd_empenhada': item.qtd_empenhada_objeto(),
            'qtd_a_empenhar': item.qtd_a_empenhar_objeto(),
            'valor_empenhado': item.valor_empenhado_objeto(),
            'valor_a_empenhar': item.valor_a_empenhar_objeto(),

            #valores
            'valor_unitario': item.valor_unitario,
            'valor_total': item.valor_total(),

            #observações
            'observacoes': item.observacoes_gerais if item.observacoes_gerais else '',
        }
        return JsonResponse(data)
    except ContratosObjetos.DoesNotExist:
        return JsonResponse({'error': 'Objeto não encontrado.'}, status=404)

def contrato_objeto_salvar(request, id_objeto=None):
    if id_objeto:
        objeto = ContratosObjetos.objects.get(id=id_objeto)
    else:
        objeto = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if objeto:
            objeto_form = ContratosObjetosForm(request.POST, instance=objeto)
            novo_objeto = False
        else:
            objeto_form = ContratosObjetosForm(request.POST)
            novo_objeto = True
        
        #Verificar se houve alteração no formulário
        if not objeto_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Instanciar: Produto
        produto_id = request.POST.get('ctobjeto_produto_hidden')
        print('PRODUTO: ', produto_id)
        produto_instance = ProdutosFarmaceuticos.objects.get(id=produto_id)
    
        #Instanciar: Fornecedor
        contrato_id = request.POST.get('contrato')
        print('CONTRATO: ', contrato_id)
        contrato_instance =  Contratos.objects.get(id=contrato_id)

        #Instanciar: ARP Item
        arp_item_id = request.POST.get('arp_item')
        if arp_item_id:
            print('ARP Item: ', arp_item_id)
            arp_item_instance = ContratosArpsItens.objects.get(id=arp_item_id)
            modificacoes_post['arp_item'] = arp_item_instance

        #Valor unitario
        valor_unitario = request.POST.get('valor_unitario')
        valor_unitario = valor_unitario.replace('R$', '').replace('\xa0', '').replace(',', '.')

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['produto'] = produto_instance
        modificacoes_post['contrato'] = contrato_instance
        modificacoes_post['valor_unitario'] = valor_unitario
        

        #Criar o formulário com os dados atualizados
        objeto_form = ContratosObjetosForm(modificacoes_post, instance=objeto_form.instance)
        
        #salvar
        if objeto_form.is_valid():
            #Salvar o produto
            objeto = objeto_form.save(commit=False)
            objeto.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_id = objeto.id
            log_registro_usuario = objeto.usuario_registro.dp_nome_completo
            log_registro_data = objeto.registro_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_atualizacao_usuario = objeto.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = objeto.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = objeto.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Contratos_Objetos",
                model='ContratosObjetos',
                model_id=objeto.id,
                item_id=0,
                item_descricao="Salvar edição de Objeto de Contrato.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Objeto (ID {objeto.id}, Número do Contrato: {objeto.contrato.numero_contrato}, Produto: {objeto.produto.produto}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_id': log_id,
                    'log_registro_usuario': log_registro_usuario,
                    'log_registro_data': log_registro_data,
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': novo_objeto,
                    'redirect_url': reverse('contrato_ficha', args=[objeto.id]),
                })
        else:
            print("Erro formulário Objeto do Contrato")
            print(objeto_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def contrato_objeto_delete(request, id_objeto=None):    
    try:
        objeto = ContratosObjetos.objects.get(id=id_objeto)
        objeto.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Objetos",
            model='ContratosObjetos',
            model_id=objeto.id,
            item_id=0,
            item_descricao="Deleção de Objeto.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Objeto (ID {objeto.id}, Nº Item: {objeto.numero_item}, Produto: {objeto.produto.produto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Objeto do Contrato deletado com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Objeto do Contrato não encontrada."
            })

def buscar_objeto(request, id_objeto=None):
    objeto = ContratosObjetos.objects.get(id=id_objeto)
    objeto_id = objeto.id
    objeto_numero_item = objeto.numero_item
    objeto_produto = str(objeto.produto.produto)
    objeto_fator_embalagem = objeto.fator_embalagem
    objeto_valor_unitario = objeto.valor_unitario

    arp_item_id = objeto.arp_item_id
    arp_item_saldo = 0
    if (arp_item_id):
        arp_item = ContratosArpsItens.objects.get(id=arp_item_id)
        arp_item_saldo = arp_item.qtd_saldo()

    objeto_dados = {
            'objeto_id': objeto_id,
            'objeto_numero_item': objeto_numero_item,
            'objeto_produto': objeto_produto,
            'objeto_fator_embalagem': objeto_fator_embalagem,
            'objeto_valor_unitario': objeto_valor_unitario,
            'arp_item_saldo': arp_item_saldo,

        }
    return JsonResponse({'objeto': objeto_dados})




#CONTRATOS/PARCELAS
def contrato_parcela_salvar(request, id_parcela=None):
    if id_parcela:
        parcela = ContratosParcelas.objects.get(id=id_parcela)
    else:
        parcela = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if parcela:
            parcela_form = ContratosParcelasForm(request.POST, instance=parcela)
            nova_parcela = False
        else:
            parcela_form = ContratosParcelasForm(request.POST)
            nova_parcela = True
        
        #Verificar se houve alteração no formulário
        if not parcela_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Objeto
        objeto_id = request.POST.get('parcela_objeto_hidden')
        objeto_instance = ContratosObjetos.objects.get(id=objeto_id)
        
        #Contrato
        contrato_id = objeto_instance.contrato.id
        contrato_instance = Contratos.objects.get(id=contrato_id)

        #Qtd Contratada
        qtd_contratada = request.POST.get('qtd_contratada')
        qtd_contratada = float(qtd_contratada.replace('.', ''))

        #Qtd Contratada
        qtd_doada = request.POST.get('qtd_doada')
        qtd_doada = float(qtd_doada.replace('.', ''))

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['objeto'] = objeto_instance
        modificacoes_post['contrato'] = contrato_instance
        modificacoes_post['qtd_contratada'] = qtd_contratada
        modificacoes_post['qtd_doada'] = qtd_doada

        #Criar o formulário com os dados atualizados
        parcela_form = ContratosParcelasForm(modificacoes_post, instance=parcela_form.instance)

        #salvar
        if parcela_form.is_valid():
            #Salvar o produto
            parcela = parcela_form.save(commit=False)
            parcela.save(current_user=request.user.usuario_relacionado)

            #id do contrato e da parcela
            contrado_id = parcela.objeto.contrato.id
            parcela_id = parcela.id

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Contratos_Parcelas",
                model='ContratosParcelas',
                model_id=parcela.id,
                item_id=0,
                item_descricao="Salvar edição de Parcela de Contrato.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou a Parcela (ID {parcela.id}, Número do Contrato: {parcela.objeto.contrato.numero_contrato}, Produto: {parcela.objeto.produto.produto}) em {current_date_str}."
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': nova_parcela,
                    'redirect_url': reverse('contrato_ficha', args=[contrado_id]),
                    'parcela_id': parcela_id,
                })
        else:
            print("Erro formulário Parcela do Contrato")
            print(parcela_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def contrato_parcela_modal(request, id_parcela=None):
    try:
        item = ContratosParcelas.objects.get(id=id_parcela)
        produto = item.objeto.produto.produto
        if item.objeto.arp_item:
            saldo_arp = item.objeto.arp_item.qtd_saldo()
        else:
            saldo_arp = 'NSA'
        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'numero_parcela': item.numero_parcela,
            'numero_item': item.objeto.numero_item,
            'produto': produto,
            'objeto': item.objeto.id,
            'fator_embalagem': item.objeto.fator_embalagem,
            'qtd_contratada': item.qtd_contratada,
            'qtd_doada': item.qtd_doada,
            'qtd_entregue': item.qtd_entregue(),
            'qtd_a_entregar': item.qtd_a_entregar(),
            'valor_unitario': item.valor_unitario(),
            'valor_total': item.valor_total(),
            'data_previsao_entrega': item.data_previsao_entrega,
            'data_ultima_entrega': item.data_ultima_entrega(),
            'observacoes': item.observacoes_gerais,

            #empenho
            'qtd_empenhada': item.qtd_empenhada(),
            'qtd_a_empenhar': item.qtd_a_empenhar(),
            'valor_empenhado': item.valor_empenhado(),
            'valor_a_empenhar': item.valor_a_empenhar(),
            'empenho_percentual': item.empenho_percentual(),
            'saldo_arp': saldo_arp,
            

        }
        return JsonResponse(data)
    except ContratosParcelas.DoesNotExist:
        return JsonResponse({'error': 'Parcela não encontrada.'}, status=404)

def buscar_parcela(request, id_parcela=None):
    parcela = ContratosParcelas.objects.get(id=id_parcela)
    numero_contrato = parcela.contrato.numero_contrato
    numero_item = parcela.objeto.numero_item
    numero_parcela = parcela.numero_parcela
    contrato_id = parcela.contrato.id
    parcela_id = parcela.id
    produto = parcela.objeto.produto.produto
    fator_embalagem = parcela.objeto.fator_embalagem
    valor_unitario = parcela.objeto.valor_unitario
    qtd_a_empenhar = parcela.qtd_a_empenhar()
    valor_a_empenhar = parcela.valor_a_empenhar()
    qtd_a_entregar = parcela.qtd_a_entregar()
    qtd_empenhada = parcela.qtd_empenhada()
    qtd_doada = parcela.qtd_doada
    total_entregue = parcela.qtd_entregue()
    parcela_dados = {
            'numero_contrato': numero_contrato,
            'numero_item': numero_item,
            'numero_parcela': numero_parcela,
            'contrato_id': contrato_id,
            'parcela_id': parcela_id,
            'produto': produto,
            'fator_embalagem': fator_embalagem,
            'valor_unitario': valor_unitario,
            'qtd_a_empenhar': qtd_a_empenhar,
            'valor_a_empenhar': valor_a_empenhar,
            'qtd_a_entregar': qtd_a_entregar,
            'qtd_empenhada': qtd_empenhada,
            'qtd_doada': qtd_doada,
            'total_entregue': total_entregue,
        }
    return JsonResponse({'parcela': parcela_dados})

def buscar_parcelas(request, id_contrato=None):
    parcelas = ContratosParcelas.objects.filter(contrato=id_contrato).order_by('objeto__numero_item', 'numero_parcela')
    parcelas_list = []
    for parcela in parcelas:
        parcela_qtd_a_empenhar = parcela.qtd_a_empenhar()
        parcela_detalhe = f"Item: {parcela.objeto.numero_item} - Parcela: {parcela.numero_parcela} - {parcela.objeto.produto.produto}"
        parcelas_list.append({'id': parcela.id, 'detalhe': parcela_detalhe, 'qtd_a_empenhar': parcela_qtd_a_empenhar})
    return JsonResponse({'parcelas': parcelas_list})

def contrato_parcela_delete(request, id_entrega=None):    
    try:
        parcela = ContratosParcelas.objects.get(id=id_entrega)
        parcela.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Entregas",
            model='ContratosEntregas',
            model_id=parcela.id,
            item_id=0,
            item_descricao="Deleção da Parcela de Contrato.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a Parcela (ID {parcela.id}, Nº Parcela: {parcela.numero_parcela}, Contrato: {parcela.contrato.numero_contrato}, Produto: {parcela.objeto.produto.produto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Parcela do Contrato deletada com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Parcela do Contrato não encontrada."
            })




#CONTRATOS/ENTREGAS
def contrato_entrega_salvar(request, id_entrega=None):
    if id_entrega:
        entrega = ContratosEntregas.objects.get(id=id_entrega)
    else:
        entrega = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if entrega:
            entrega_form = ContratosEntregasForm(request.POST, instance=entrega)
            nova_entrega = False
        else:
            entrega_form = ContratosEntregasForm(request.POST)
            nova_entrega = True
        
        #Verificar se houve alteração no formulário
        if not entrega_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Parcela
        parcela_id = request.POST.get('id_entrega_parcela_hidden')
        parcela_instance = ContratosParcelas.objects.get(id=parcela_id)
        
        #Contrato
        contrato_id = parcela_instance.contrato.id
        contrato_instance = Contratos.objects.get(id=contrato_id)

        #Qtd Contratada
        qtd_entregue = request.POST.get('qtd_entregue')
        qtd_entregue = float(qtd_entregue.replace('.', ''))

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['parcela'] = parcela_instance
        modificacoes_post['contrato'] = contrato_instance
        modificacoes_post['qtd_entregue'] = qtd_entregue

        #Criar o formulário com os dados atualizados
        entrega_form = ContratosEntregasForm(modificacoes_post, instance=entrega_form.instance)

        print('Observações: ', request.POST.get('observacoes_gerais'))

        #salvar
        if entrega_form.is_valid():
            #Salvar o produto
            entrega = entrega_form.save(commit=False)
            entrega.save(current_user=request.user.usuario_relacionado)

            #id do contrato e da parcela
            contrado_id = entrega.parcela.contrato.id
            parcela_id = entrega.parcela.id
            entrega_id = entrega.id

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            observacoes = (
                f"Usuário {request.user.username} salvou a Entrega "
                f"(ID {entrega.id}, Nº da Entrega: {entrega.numero_entrega}, "
                f"Parcela: {entrega.parcela}, "
                f"Contrato: {entrega.contrato}, "
                f"Fornecedor: {entrega.contrato.fornecedor}, "
                f"Produto: {entrega.parcela.objeto.produto.produto}) em {current_date_str}."
            )
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Contratos_Entregas",
                model='ContratosEntregas',
                model_id=entrega.id,
                item_id=0,
                item_descricao="Salvar edição da Entrega de Contrato.",
                acao="Salvar",
                observacoes=observacoes
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': nova_entrega,
                    'entrega_id': entrega_id,
                })
        else:
            print("Erro formulário Parcela do Contrato")
            print(entrega_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def contrato_entrega_modal(request, id_entrega=None):
    try:
        item = ContratosEntregas.objects.get(id=id_entrega)
        produto = item.parcela.objeto.produto.produto
        contrato = item.contrato.id
        parcela = item.parcela.id
        qtd_a_entregar = item.parcela.qtd_a_entregar()
        data = {
            'entrega_id': item.id,
            'entrega_log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'entrega_log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'entrega_log_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'entrega_log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'entrega_log_edicoes': item.log_n_edicoes,
            
            'id_entrega_item': item.id,
            'id_entrega_parcela': item.parcela.numero_parcela,
            'id_entrega_numero_entrega': item.numero_entrega,
            'id_entrega_produto': produto,
            'id_entrega_qtd_entregue': item.qtd_entregue,
            'id_data_entrega': item.data_entrega,
            'id_entrega_local_entrega': item.local_entrega,
            'id_entrega_notas_recebidas': item.notas_recebidas,
            'id_entrega_notas_status': item.notas_status,
            'id_entrega_notas_pagamentos': item.notas_pagamentos,
            'observacoes_gerais': item.observacoes_gerais,

            'id_entrega_contrato_hidden': contrato,
            'id_entrega_parcela_hidden': parcela,
            'id_qtd_a_entregar_hidden': qtd_a_entregar, 

            'qtd_empenhada': item.parcela.qtd_empenhada(),
            'qtd_doada': item.parcela.qtd_doada,
            'total_entregue': item.parcela.qtd_entregue(),
        }
        return JsonResponse(data)
    except ContratosEntregas.DoesNotExist:
        return JsonResponse({'error': 'Entrega não encontrada.'}, status=404)

def contrato_entrega_delete(request, id_entrega=None):    
    try:
        entrega = ContratosEntregas.objects.get(id=id_entrega)
        entrega.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Entregas",
            model='ContratosEntregas',
            model_id=entrega.id,
            item_id=0,
            item_descricao="Deleção de Entrega de Contrato.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a Entrega (ID {entrega.id}, Nº Parcela: {entrega.parcela.numero_parcela}, Contrato: {entrega.contrato.numero_contrato}, Produto: {entrega.parcela.objeto.produto.produto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Entrega do Contrato deletada com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Entrega do Contrato não encontrada."
            })




#CONTRATOS/FISCAIS
def contrato_fiscal_salvar(request, id_fiscal=None):
    if id_fiscal:
        fiscal = ContratosFiscais.objects.get(id=id_fiscal)
    else:
        fiscal = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if fiscal:
            fiscal_form = ContratosFiscaisForm(request.POST, instance=fiscal)
            novo_fiscal = False
        else:
            fiscal_form = ContratosFiscaisForm(request.POST)
            novo_fiscal = True
        
        #Verificar se houve alteração no formulário
        if not fiscal_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })
        
        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Usuario
        outro_fiscal = request.POST.get('fiscal_outro')
        if outro_fiscal == '':
            print('teste1')
            fiscal_id = request.POST.get('fiscal_usuario_hidden')
            fiscal_instance = Usuario.objects.get(id=fiscal_id)
  
            #Atualizar os valores no mutable_post
            modificacoes_post['fiscal'] = fiscal_instance
        
        #Contrato
        contrato_id = request.POST.get('id_fiscal_contrato_hidden')
        contrato_instance = Contratos.objects.get(id=contrato_id)

        #Atualizar os valores no mutable_post
        modificacoes_post['contrato'] = contrato_instance

        #Criar o formulário com os dados atualizados
        fiscal_form = ContratosFiscaisForm(modificacoes_post, instance=fiscal_form.instance)
        
        #salvar
        if fiscal_form.is_valid():
            #Salvar o produto
            fiscal = fiscal_form.save(commit=False)
            fiscal.save(current_user=request.user.usuario_relacionado)
            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            observacoes = (
                f"Usuário {request.user.username} salvou o Fiscal do Contrato "
                f"(ID {fiscal.id}, Nome: {fiscal.fiscal_nome()}, "
                f"Contrato: (ID: {fiscal.contrato.id}, Número: {fiscal.contrato.numero_contrato}) "
                f"em {current_date_str}."
            )
            
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Contratos_Fiscais",
                model='ContratosFiscais',
                model_id=fiscal.id,
                item_id=0,
                item_descricao="Salvar edição de Fiscal de Contrato.",
                acao="Salvar",
                observacoes=observacoes
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': novo_fiscal,
                    'fiscal_id': fiscal.id,
                })
        else:
            print("Erro formulário Fiscal do Contrato")
            print(fiscal_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def contrato_fiscal_modal(request, id_fiscal=None):
    try:
        item = ContratosFiscais.objects.get(id=id_fiscal)
        fiscal = None
        fiscal_outro = item.fiscal_outro
        if fiscal_outro == '':
            fiscal = item.fiscal.id
            fiscal_outro_checkbox = False
        else:
            fiscal_outro_checkbox = True
        contrato = item.contrato.id
        maior_data_fim = ContratosFiscais.objects.filter(contrato=contrato).aggregate(Max('data_fim'))['data_fim__max']

        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'log_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            
            'id_fiscal': fiscal,
            'id_fiscal_outro_checkbox': fiscal_outro_checkbox,
            'id_fiscal_outro': fiscal_outro,
            'id_fiscal_status': item.status,
            'id_fiscal_documento_sei': item.documento_sei,

            'id_fiscal_data_inicio': item.data_inicio,
            'id_fiscal_data_fim': item.data_fim,
            'id_fiscal_observacoes': item.observacoes_gerais,

            'id_fiscal_contrato_hidden': contrato,

            'id_fiscal_data_fim_maior_hidden': maior_data_fim,
        }
        return JsonResponse(data)
    except ContratosEntregas.DoesNotExist:
        return JsonResponse({'error': 'Fiscal não encontrado.'}, status=404)

def contrato_fiscal_delete(request, id_fiscal=None):
    try:
        fiscal = ContratosFiscais.objects.get(id=id_fiscal)
        fiscal.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Fiscais",
            model='ContratosFiscais',
            model_id=fiscal.id,
            item_id=0,
            item_descricao="Deleção de Fiscal de Contrato.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Fiscal do Contrato (ID {fiscal.id}, Nome: {fiscal.fiscal_nome}, Contrato: {fiscal.contrato.numero_contrato}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Fiscal do Contrato deletado com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Fiscal do Contrato não encontrado."
            })




#ANOTACOES DO CONTRATO
def contrato_anotacoes(request, id_contrato):
    return render(request, 'contratos/contrato_anotacoes.html')




#ARPs
def arps(request):
    tab_arps = ContratosArps.objects.all().filter(del_status=False).order_by('-data_publicacao')
    denominacoes = DenominacoesGenericas.objects.values_list('id', 'denominacao')
    fornecedores = Fornecedores.objects.values_list('nome_fantasia', flat=True).distinct().order_by('nome_fantasia')
    unidades_daf = [item for item in UNIDADE_DAF if item[0] not in ['cofisc', 'gabinete']]
    conteudo = {
        'lista_status': STATUS_ARP,
        'lista_unidadesdaf': unidades_daf,
        'lista_denominacoes': denominacoes,
        'lista_fornecedores': fornecedores,
        'tab_arps': tab_arps,
    }
    return render(request, 'contratos/arps.html', conteudo)

def arp_ficha(request, arp_id=None):
    if arp_id:
        try:
            arp = ContratosArps.objects.get(id=arp_id)
        except ContratosArps.DoesNotExist:
            messages.error(request, "ARP não encontrada.")
            return redirect('arps')
    else:
        arp = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if arp:
            arp_form = ContratosArpsForm(request.POST, instance=arp)
            nova_arp = False
        else:
            arp_form = ContratosArpsForm(request.POST)
            nova_arp = True

        #Verificar se houve alteração no formulário
        if not arp_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })
        
        #Passar o objeto Denominação Genérica
        denominacao_id = request.POST.get('denominacao')
        if denominacao_id == None:
            denominacao_id = request.POST.get('arp_denominacao_hidden')
        denominacao_instance = DenominacoesGenericas.objects.get(id=denominacao_id)
    
        #Passar o objeto Fornecedor
        fornecedor_id = request.POST.get('arp_fornecedor_hidden')
        print('Fornecedor: ', fornecedor_id)
        fornecedor_instance =  Fornecedores.objects.get(id=fornecedor_id)
        
        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['denominacao'] = denominacao_instance
        modificacoes_post['fornecedor'] = fornecedor_instance

        #Criar o formulário com os dados atualizados
        arp_form = ContratosArpsForm(modificacoes_post, instance=arp_form.instance)

        #salvar
        if arp_form.is_valid():
            #Salvar o produto
            arp = arp_form.save(commit=False)
            arp.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_id = arp.id
            log_registro_usuario = arp.usuario_registro.dp_nome_completo
            log_registro_data = arp.registro_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_atualizacao_usuario = arp.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = arp.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = arp.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_ARPs",
                model='ContratosARPs',
                model_id=arp.id,
                item_id=0,
                item_descricao="Salvar edição de ARP.",
                acao="Salvar",
                #observacoes=f"Usuário {request.user.username} salvou a ARP (ID {arp.id}, Número: {arp.topico}, Denominação: {arp.denominacao.denominacao}, Fornecedor: {arp.fornecedor.nome_fantasia}) em {current_date_str}."
                observacoes=f"Usuário {request.user.username} salvou a ARP (ID {arp.id}, Número: {arp.numero_arp}, Denominação: {arp.denominacao.denominacao}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_id': log_id,
                    'log_registro_usuario': log_registro_usuario,
                    'log_registro_data': log_registro_data,
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': nova_arp,
                    'redirect_url': reverse('arp_ficha', args=[arp.id]),
                })
        else:
            print("Erro formulário ARP")
            print(arp_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })
            
    #Form ARP
    tab_itens_arp = None
    if arp:
        form = ContratosArpsForm(instance=arp)
        
        #Itens da ARP
        tab_itens_arp = ContratosArpsItens.objects.filter(del_status=False, arp_id=arp.id).order_by('numero_item')
        
        #Calculando o valor total para cada item
        tab_itens_arp = tab_itens_arp.annotate(
            valor_total=Case(
                When(valor_unit_reequilibrio_bool=True, then=F('valor_unit_reequilibrio') * F('qtd_registrada')),
                default=F('valor_unit_homologado') * F('qtd_registrada'),
                output_field=models.FloatField()
                )
            )
    else:
        form = ContratosArpsForm()

    #Form Item da ARP
    form_item = ContratosArpsItensForm()

    return render(request, 'contratos/arp_ficha.html', {
        'YES_NO': YES_NO,
        'TIPO_COTA': TIPO_COTA,
        'form': form,
        'form_item': form_item,
        'arp': arp,
        'tab_itens_arp': tab_itens_arp,
    })

def arp_delete(request, arp_id=None):   
    try:
        arp = ContratosArps.objects.get(id=arp_id)
        arp.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_ARPs",
            model='ContratosARPs',
            model_id=arp.id,
            item_id=0,
            item_descricao="Deleção de ARP.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a ARP (ID {arp.id}, Número: {arp.numero_arp}, Denominação: {arp.denominacao.denominacao}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "ARP deletada com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "ARP não encontrada."
            })

def arp_buscar_produtos(request, denominacao=None):
    produtos = ProdutosFarmaceuticos.get_produtos_por_denominacao(denominacao)
    return JsonResponse(produtos, safe=False)

def arp_filtrar(request):
    status_arp = request.GET.get('status_arp', None)
    unidade_daf = request.GET.get('unidade_daf', None)
    denominacao = request.GET.get('denominacao', None)
    fornecedor = request.GET.get('fornecedor', None)

    filters = {}
    filters['del_status'] = False
    if status_arp:
        filters['status'] = status_arp
    if unidade_daf:
        filters['unidade_daf'] = unidade_daf
    if denominacao:
        filters['denominacao_id'] = denominacao
    if fornecedor:
        filters['fornecedor__nome_fantasia__icontains'] = fornecedor
    
    tab_arps = ContratosArps.objects.filter(**filters).order_by('-data_publicacao')
    total_arps = tab_arps.count()

    page = int(request.GET.get('page', 1))
    paginator = Paginator(tab_arps, 100)  # Mostra 100 faqs por página
    try:
        arp_paginados = paginator.page(page)
    except EmptyPage:
        arp_paginados = paginator.page(paginator.num_pages)

    data = []
    for arp in arp_paginados.object_list:
        valor_total = arp.valor_total_arp()
        arp_data = {
            'id': arp.id,
            'status': arp.status,
            'unidade_daf': arp.unidade_daf,
            'numero_processo_sei': arp.numero_processo_sei,
            'numero_documento_sei': arp.numero_documento_sei,
            'data_publicacao': arp.data_publicacao.strftime('%d/%m/%Y') if arp.data_publicacao else '',
            'data_vigencia': arp.data_vigencia.strftime('%d/%m/%Y') if arp.data_publicacao else '',
            'prazo_vigencia': arp.prazo_vigencia if arp.data_publicacao else '',
            'denominacao': arp.denominacao.denominacao,
            'fornecedor': arp.fornecedor.nome_fantasia,
            'valor_total_arp': valor_total,
        }
        data.append(arp_data)

    return JsonResponse({
        'data': data,
        'total_arps': total_arps,
        'has_next': arp_paginados.has_next(),
        'has_previous': arp_paginados.has_previous(),
        'current_page': page
    })

def arp_exportar(request):
    print("Exportar ARPs")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        status_arp = data.get('status_arp')
        unidade_daf = data.get('unidade_daf')
        denominacao = data.get('denominacao')
        fornecedor = data.get('fornecedor')

        filters = {}
        filters['del_status'] = False
        if status_arp:
            filters['status'] = status_arp
        if unidade_daf:
            filters['unidade_daf'] = unidade_daf
        if denominacao:
            filters['denominacao_id'] = denominacao
        if fornecedor:
            filters['fornecedor__nome_fantasia__icontains'] = fornecedor
        
        arps = ContratosArps.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "arps"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização', 'N Edições', 
        'Unidade DAF', 'Processo SEI', 'Documento SEI', 'ARP', 'Status', 'Data Publicacao',
        'Denominacao', 'Fornecedor',
        'Observações Gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 20

        # Adicionar dados da tabela
        for row_num, arp in enumerate(arps, 2):
            registro_data = arp.registro_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
            ult_atual_data = arp.ult_atual_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
            data_publicacao = arp.data_publicacao
            if data_publicacao:
                data_publicacao.strftime('%d/%m/%Y')
            denominacao = str(arp.denominacao)
            fornecedor = str(arp.fornecedor)

            ws.cell(row=row_num, column=1, value=arp.id)
            ws.cell(row=row_num, column=2, value=str(arp.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(arp.usuario_atualizacao.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=arp.log_n_edicoes)
            ws.cell(row=row_num, column=7, value=arp.unidade_daf)
            ws.cell(row=row_num, column=8, value=arp.numero_processo_sei)
            ws.cell(row=row_num, column=9, value=arp.numero_documento_sei)
            ws.cell(row=row_num, column=10, value=arp.numero_arp)
            ws.cell(row=row_num, column=11, value=arp.status)
            ws.cell(row=row_num, column=12, value=data_publicacao)
            ws.cell(row=row_num, column=13, value=denominacao)
            ws.cell(row=row_num, column=14, value=fornecedor)
            ws.cell(row=row_num, column=15, value=arp.observacoes_gerais)
            ws.cell(row=row_num, column=16, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_ARPs",
            model='ContratosARPs',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de ARPs",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou lista de ARPs em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_fornecedores.xlsx"'
        response.write(output.getvalue())
        return response

def arp_buscar_dados_sei(request, id_arp=None):
    arp = ContratosArps.objects.filter(id=id_arp)
    arp_list = list(arp.values('id', 'numero_processo_sei', 'lei_licitacao'))
    return JsonResponse({'arp': arp_list})




#ITENS DAS ARPS
def arp_item_ficha(request, arp_item_id=None):
    if arp_item_id:
        try:
            item_arp = ContratosArpsItens.objects.get(id=arp_item_id)
        except ContratosArpsItens.DoesNotExist:
            messages.error(request, "Item da ARP não encontrada.")
            return redirect('arps')
    else:
        item_arp = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if item_arp:
            item_arp_form = ContratosArpsItensForm(request.POST, instance=item_arp)
            novo_item_arp = False
        else:
            item_arp_form = ContratosArpsItensForm(request.POST)
            novo_item_arp = True

        #Verificar se houve alteração no formulário
        if not item_arp_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Passar o objeto ARP
        arp_id = request.POST.get('id_arp')
        arp_instance = ContratosArps.objects.get(id=arp_id)

        #Passar o objeto Produto Farmacêutico
        produto_id = request.POST.get('produto')
        produto_instance = ProdutosFarmaceuticos.objects.get(id=produto_id)

        #valor unitario homologado
        valor_homologado_str  = request.POST.get('valor_unit_homologado')
        valor_homologado_str = valor_homologado_str.replace('R$', '').replace('.', '')
        valor_homologado_str = valor_homologado_str.replace(',', '.')
        valor_unit = float(valor_homologado_str)

        #valor unitario reequilibrio
        valor_reequilibrio_str  = request.POST.get('valor_unit_reequilibrio')
        if valor_reequilibrio_str != "":
            valor_reequilibrio_str = valor_reequilibrio_str.replace('R$', '').replace('.', '')
            valor_reequilibrio_str = valor_reequilibrio_str.replace(',', '.')
            valor_reequilibrio = float(valor_reequilibrio_str)
            modificacoes_post['valor_unit_reequilibrio'] = valor_reequilibrio

        #qtd registrada
        qtd_registrada_str = request.POST.get('qtd_registrada')
        qtd_registrada_str = qtd_registrada_str.replace('.', '')
        qtd_registrada_int = int(qtd_registrada_str)

        #Atualizar os valores no mutable_post
        modificacoes_post['arp'] = arp_instance
        modificacoes_post['produto'] = produto_instance
        modificacoes_post['valor_unit_homologado'] = valor_unit
        modificacoes_post['qtd_registrada'] = qtd_registrada_int

        #Criar o formulário com os dados atualizados
        item_arp_form = ContratosArpsItensForm(modificacoes_post, instance=item_arp_form.instance)

        #salvar
        if item_arp_form.is_valid():
            #Salvar o produto
            item_arp = item_arp_form.save(commit=False)
            item_arp.save(current_user=request.user.usuario_relacionado)
            
            #logs
            item_arp_id = item_arp.id

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_ARPs_Itens",
                model='ContratosARPsItens',
                model_id=item_arp.id,
                item_id=0,
                item_descricao="Salvar edição de Item da ARP.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Item da ARP (ID {item_arp.id}, Nº Item: {item_arp.numero_item}, Nº ARP: {item_arp.arp.numero_arp}, Produto: {item_arp.produto.produto}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'item_arp_id': item_arp_id,
                    'novo': novo_item_arp,
                    'redirect_url': reverse('arp_ficha', args=[item_arp.arp_id]),
                })
        else:
            print("Erro formulário Item da ARP")
            print(item_arp_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })
            
    #Form ARP
    if item_arp:
        form_item = ContratosArpsItensForm(instance=item_arp)
    else:
        form_item = ContratosArpsItensForm()

    return render(request, 'contratos/arp_ficha.html', {
        'YES_NO': YES_NO,
        'TIPO_COTA': TIPO_COTA,
        'form_item': form_item,
        'item_arp': item_arp,
    })

def arp_item_formulario(request, arp_item_id=None):
    try:
        item = ContratosArpsItens.objects.get(id=arp_item_id)
        produto_id = item.produto_id
        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'numero_item': item.numero_item,
            'tipo_cota': item.tipo_cota,
            'empate_ficto': item.empate_ficto,
            'produto': produto_id,
            'valor_unit_homologado': item.valor_unit_homologado,
            'valor_unit_reequilibrio_bool': item.valor_unit_reequilibrio_bool,
            'valor_unit_reequilibrio': item.valor_unit_reequilibrio,
            'qtd_registrada': item.qtd_registrada,
            'observacoes': item.observacoes_gerais if item.observacoes_gerais else '',
            'contratos': item.contratos(),
            'qtd_contratada': item.qtd_contratada(),
            'valor_contratado': item.valor_contratado(),
            'saldo_quantidade': item.qtd_saldo(),
            'saldo_valor': item.valor_saldo(),
            'saldo_percentual': item.qtd_saldo_percentual(),
        }
        return JsonResponse(data)
    except ContratosArpsItens.DoesNotExist:
        return JsonResponse({'error': 'Item da ARP não encontrada'}, status=404)

def arp_item_delete(request, arp_item_id=None):   
    try:
        item_arp = ContratosArpsItens.objects.get(id=arp_item_id)
        item_arp.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_ARPs_Itens",
            model='ContratosARPsItens',
            model_id=item_arp.id,
            item_id=0,
            item_descricao="Deleção de Itens da ARP.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Item da ARP (ID {item_arp.id}, Nº Item: {item_arp.numero_item}, Produto: {item_arp.produto.produto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Item da ARP deletada com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Item da ARP não encontrada."
            })




#EMPENHOS
def empenhos(request):
    tabEmpenhos = Empenhos.objects.filter(del_status=False).annotate(
            data_empenho_ordenada=Case(
                When(data_empenho__isnull=True, then=Value(0)),
                default=Value(1)
            )
        ).order_by('-data_empenho_ordenada', '-data_empenho')
    conteudo = {
        'tabEmpenhos': tabEmpenhos,
    }
    return render(request, 'contratos/empenhos.html', conteudo)

def empenho_ficha(request, id_empenho=None):
    if id_empenho:
        empenho = Empenhos.objects.get(id=id_empenho)
    else:
        empenho = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if empenho:
            empenho_form = EmpenhoForm(request.POST, instance=empenho)
            novo_empenho = False
        else:
            empenho_form = EmpenhoForm(request.POST)
            novo_empenho = True
        
        #Verificar se houve alteração no formulário
        if not empenho_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })
        
        #salvar
        if empenho_form.is_valid():
            #Salvar o produto
            empenho = empenho_form.save(commit=False)
            empenho.save(current_user=request.user.usuario_relacionado)
            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            observacoes = (
                f"Usuário {request.user.username} salvou o Empenho Orçamentário "
                f"(ID {empenho.id}) "
                f"em {current_date_str}."
            )
            
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Empenhos",
                model='Empenhos',
                model_id=empenho.id,
                item_id=0,
                item_descricao="Salvar edição de Empenho Orçamentário.",
                acao="Salvar",
                observacoes=observacoes
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': novo_empenho,
                    'empenho_id': empenho.id,
                })
        else:
            print("Erro formulário Fiscal do Contrato")
            print(empenho_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })
    
    list_contratos = []
    if empenho:
        empenho_form = EmpenhoForm(instance=empenho)
        empenho_unidade_daf = empenho.unidade_daf
        
        tab_contratos = Contratos.objects.filter(unidade_daf=empenho_unidade_daf, del_status=False, status='vigente')
        if tab_contratos.count()>0:
            for contrato in tab_contratos:
                contrato_id = contrato.id
                item_contrato = f'[ID: {contrato.id}] Contrato Nº {contrato.numero_contrato} - {contrato.denominacao.denominacao} - {contrato.fornecedor.nome_fantasia}'
                list_contratos.append((contrato_id, item_contrato))
        
        tab_empenho_itens = EmpenhosItens.objects.filter(del_status=False, empenho=empenho.id)
    else:
        empenho_form = EmpenhoForm()
        tab_empenho_itens = None

    form_empenho_item = EmpenhosItensForm()

    conteudo = {
        'form': empenho_form,
        'form_empenho_item': form_empenho_item,
        'empenho': empenho,
        'tab_empenho_itens': tab_empenho_itens,
        'list_contratos': list_contratos,
    }
    return render(request, 'contratos/empenho_ficha.html', conteudo)

def empenho_deletar(request, id_empenho=None):   
    try:
        empenho = Empenhos.objects.get(id=id_empenho)
        empenho.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Empenhos",
            model='Empenhos',
            model_id=empenho.id,
            item_id=0,
            item_descricao="Deleção do Empenho.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Empenho (ID {empenho.id}, Nº Empenho: {empenho.numero_empenho}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Empenho deletado com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Empenho não encontrado."
            })




#ITENS DO EMPENHO
def item_empenho_salvar(request, item_empenho_id=None):
    if item_empenho_id:
        try:
            item_empenho = EmpenhosItens.objects.get(id=item_empenho_id)
        except EmpenhosItens.DoesNotExist:
            messages.error(request, "Item do Empenho não encontrado.")
            return redirect('empenhos')
    else:
        item_empenho = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if item_empenho:
            item_empenho_form = EmpenhosItensForm(request.POST, instance=item_empenho)
            novo_item_empenho = False
        else:
            item_empenho_form = EmpenhosItensForm(request.POST)
            novo_item_empenho = True

        #Verificar se houve alteração no formulário
        if not item_empenho_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Passar o objeto Empenho
        empenho_id = request.POST.get('id_empenhoItem_empenho')
        empenho_instance = Empenhos.objects.get(id=empenho_id)

        #Passar o objeto Contrato
        parcela_id = request.POST.get('id_empenhoItem_parcela')
        parcela_instance = ContratosParcelas.objects.get(id=parcela_id)

        #valor total empenhado
        valor_empenhado_str  = request.POST.get('valor_empenhado')
        valor_empenhado_str = valor_empenhado_str.replace('R$', '').replace('.', '')
        valor_empenhado_str = valor_empenhado_str.replace(',', '.')
        valor_empenhado = float(valor_empenhado_str)

        #qtd empenhada
        qtd_empenhada_str = request.POST.get('qtd_empenhado')
        qtd_empenhada_str = qtd_empenhada_str.replace('.', '')
        qtd_empenhada = float(qtd_empenhada_str)

        #Atualizar os valores no mutable_post
        modificacoes_post['empenho'] = empenho_instance
        modificacoes_post['parcela'] = parcela_instance
        modificacoes_post['valor_empenhado'] = valor_empenhado
        modificacoes_post['qtd_empenhado'] = qtd_empenhada

        #Criar o formulário com os dados atualizados
        item_empenho_form = EmpenhosItensForm(modificacoes_post, instance=item_empenho_form.instance)

        #salvar
        if item_empenho_form.is_valid():
            #Salvar o produto
            item_empenho = item_empenho_form.save(commit=False)
            item_empenho.save(current_user=request.user.usuario_relacionado)

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Empenhos_Itens",
                model='EmpenhosItens',
                model_id=item_empenho.id,
                item_id=0,
                item_descricao="Salvar edição de Item do Empenho.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Item do Empenho (ID {item_empenho.id}, ID Parcela: {item_empenho.parcela.id}, Produto: {item_empenho.parcela.objeto.produto}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': novo_item_empenho,
                    'item_empenho_id': item_empenho.id,
                    'redirect_url': reverse('empenho_ficha', args=[item_empenho.empenho.id]),
                })
        else:
            print("Erro formulário Item da ARP")
            print(item_empenho_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def item_empenho_modal(request, item_empenho_id=None):
    try:
        item = EmpenhosItens.objects.get(id=item_empenho_id)
        produto = item.parcela.objeto.produto.produto
        data = {
            #log
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            
            #item
            'contrato': item.parcela.contrato.numero_contrato,
            'numero_item': item.parcela.objeto.numero_item,
            'numero_parcela': item.parcela.numero_parcela,
            'produto': produto,

            #parâmetros
            'fator_embalagem': item.parcela.objeto.fator_embalagem,
            'valor_unitario': item.parcela.objeto.valor_unitario,
            'qtd_a_empenhar': item.parcela.qtd_a_empenhar(),
            'valor_a_empenhar': item.parcela.valor_a_empenhar(),

            #empenho
            'qtd_embalagens': item.qtd_embalagens(),
            'qtd_empenhada': item.qtd_empenhado,
            'valor_empenhado': item.valor_empenhado,

            #observações
            'observacoes': item.observacoes_gerais,

            #campos ocultos
            'empenho_id': item.empenho.id,
            'parcela_id': item.parcela.id,

        }
        return JsonResponse(data)
    except ContratosArpsItens.DoesNotExist:
        return JsonResponse({'error': 'Item do Empenho não encontrada'}, status=404)

def item_empenho_deletar(request, item_empenho_id=None):   
    try:
        item_empenho = EmpenhosItens.objects.get(id=item_empenho_id)
        item_empenho.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Empenhos_Itens",
            model='EmpenhoItens',
            model_id=item_empenho.id,
            item_id=0,
            item_descricao="Deleção de Itens do Empenho.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Item da Empenho (ID {item_empenho.id}, Nº Empenho: {item_empenho.empenho.numero_empenho}, Nº Contrato: {item_empenho.parcela.contrato.numero_contrato}, Objeto: {item_empenho.parcela.objeto.produto.produto}, Parcela: {item_empenho.parcela.numero_parcela}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Item da Empenho deletado com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Item da Empenho não encontrado."
            })




#TEDs
def teds(request):
    return render(request, 'contratos/teds.html')




#RELATÓRIOS
def contratos_relatorios_arp(request, arp_id=None):
    arp = ContratosArps.objects.get(id=arp_id)
    arp_itens = ContratosArpsItens.objects.filter(del_status=False, arp=arp_id)
    
    #Log Relatório
    usuario_nome = request.user.usuario_relacionado.primeiro_ultimo_nome
    data_hora_atual = datetime.now()
    data_hora = data_hora_atual.strftime('%d/%m/%Y %H:%M:%S')
    
    conteudo = {
        'arp': arp,
        'arp_itens': arp_itens,
        'usuario': usuario_nome,
        'data_hora': data_hora,
    }
    return render(request, 'contratos/relatorio_arp.html', conteudo)

def contratos_relatorios_empenho(request, empenho_id=None):
    empenho = Empenhos.objects.get(id=empenho_id)
    empenho_itens = EmpenhosItens.objects.filter(del_status=False, empenho=empenho_id)
    
    #Log Relatório
    usuario_nome = request.user.usuario_relacionado.primeiro_ultimo_nome
    data_hora_atual = datetime.now()
    data_hora = data_hora_atual.strftime('%d/%m/%Y %H:%M:%S')
    
    conteudo = {
        'empenho': empenho,
        'empenho_itens': empenho_itens,
        'usuario': usuario_nome,
        'data_hora': data_hora,
    }
    return render(request, 'contratos/relatorio_empenho.html', conteudo)

def contratos_relatorios_contrato(request, contrato_id=None):
    contrato = Contratos.objects.get(id=contrato_id)
    contrato_objetos = ContratosObjetos.objects.filter(del_status=False, contrato=contrato_id)
    contrato_parcelas = ContratosParcelas.objects.filter(del_status=False, contrato=contrato_id)
    contrato_entregas = ContratosEntregas.objects.filter(del_status=False, contrato=contrato_id)
    contrato_fiscais = ContratosFiscais.objects.filter(del_status=False, contrato=contrato_id)
    
    #Log Relatório
    usuario_nome = request.user.usuario_relacionado.primeiro_ultimo_nome
    data_hora_atual = datetime.now()
    data_hora = data_hora_atual.strftime('%d/%m/%Y %H:%M:%S')
    
    conteudo = {
        'contrato': contrato,
        'contrato_objetos': contrato_objetos,
        'contrato_parcelas': contrato_parcelas,
        'contrato_entregas': contrato_entregas,
        'contrato_fiscais': contrato_fiscais,
        'usuario': usuario_nome,
        'data_hora': data_hora,
    }
    return render(request, 'contratos/relatorio_contrato.html', conteudo)
