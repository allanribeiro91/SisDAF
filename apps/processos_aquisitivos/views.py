from unidecode import unidecode
from django.core.paginator import Paginator, EmptyPage
from django.template.loader import render_to_string
from django.http import QueryDict
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import auth, messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from setup.choices import UNIDADE_DAF, UNIDADE_DAF2, MODALIDADE_AQUISICAO, STATUS_PROAQ, STATUS_FASE, FASES_EVOLUCAO_PROAQ
from apps.usuarios.models import Usuario
from apps.produtos.models import DenominacoesGenericas, ProdutosFarmaceuticos
from apps.main.models import CustomLog
from apps.processos_aquisitivos.models import (ProaqDadosGerais, ProaqProdutos, ProaqEvolucao, 
                                               PROAQ_AREA_MS, PROAQ_ETAPA, ProaqTramitacao, ProaqItens)
from apps.processos_aquisitivos.forms import (ProaqDadosGeraisForm, ProaqItensForm, 
                                              ProaqEvolucaoForm, ProaqTramitacaoForm)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils import timezone
from io import BytesIO
import json
import re
import pytz
from django.db.models import Max, OuterRef, Subquery


#timezone
tz = pytz.timezone("America/Sao_Paulo")

#Página inicial do módulo
def proaq(request):
    tab_proaqs = ProaqDadosGerais.objects.filter(del_status=False).order_by('denominacao')
    lista_denominacoes = (
        ProaqDadosGerais.objects.filter(del_status=False)
        .values_list('denominacao__id', 'denominacao__denominacao')
        .distinct()
    )
    total_processos = tab_proaqs.count()
    
    # Ordenando a lista em Python, removendo acentos e considerando maiúsculas/minúsculas
    lista_unidadesdaf = [item for item in UNIDADE_DAF if item[0] not in ['cofisc', 'gabinete', '']]

    conteudo = {
        'tab_proaqs': tab_proaqs,
        'lista_modalidades': MODALIDADE_AQUISICAO,
        'lista_denominacoes': lista_denominacoes,
        'lista_status': STATUS_PROAQ,
        'lista_evolucao': FASES_EVOLUCAO_PROAQ,
        'lista_unidadesdaf': lista_unidadesdaf,
        'total_processos': total_processos,
    }
    return render(request, 'processos_aquisitivos/proaq.html', conteudo)

@login_required
def proaq_filtro(request):
    status = request.GET.get('status', None)
    unidadeDAF = request.GET.get('unidadeDAF', None)
    faseEvolucao = request.GET.get('faseEvolucao', None)
    modalidadeAquisicao = request.GET.get('modalidadeAquisicao', None)
    denominacao_id = request.GET.get('denominacao_id', None)

    filters = {'del_status': False}
    if status:
        filters['status'] = status
    if unidadeDAF:
        filters['unidade_daf'] = unidadeDAF
    if modalidadeAquisicao:
        filters['modalidade_aquisicao'] = modalidadeAquisicao
    if denominacao_id:
        filters['denominacao_id'] = denominacao_id

    # Obter a fase mais recente de cada ProaqDadosGerais
    ultima_fase = ProaqEvolucao.objects.filter(
        proaq_id=OuterRef('pk'),
        del_status=False
    ).order_by('-fase_numero')

    # Adicionando o campo da última fase ao queryset
    tab_proaqs = ProaqDadosGerais.objects.annotate(
        ultima_fase=Subquery(ultima_fase.values('fase')[:1])
    ).filter(**filters)

    # Filtrar pela fase de evolução, se necessário
    if faseEvolucao:
        tab_proaqs = tab_proaqs.filter(ultima_fase=faseEvolucao)

    tab_proaqs = tab_proaqs.order_by('denominacao')
    total_processos = tab_proaqs.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(tab_proaqs, 100)  # Mostra 100 denominações por página
    try:
        proaqs_paginados = paginator.page(page)
    except EmptyPage:
        proaqs_paginados = paginator.page(paginator.num_pages)

    data = []

    for proaq in proaqs_paginados.object_list:
        fase_processo_resultado = proaq.fase_processo()
        proaq_dict = {
            'id': proaq.id,
            'get_status_label': proaq.get_status_label(),
            'get_unidade_daf_label': proaq.get_unidade_daf_label(),
            'get_modalidade_aquisicao_label': proaq.get_modalidade_aquisicao_label(),
            'numero_processo_sei': proaq.numero_processo_sei,
            'numero_etp': proaq.numero_etp,
            'get_denominacao_nome': proaq.get_denominacao_nome(),
            'get_usuario_nome': proaq.get_usuario_nome(),
            'total_itens': proaq.total_itens(),
            'valor_total': proaq.valor_total(),
            'fase_processo': fase_processo_resultado[0] if fase_processo_resultado else None,
            'fase_dias': fase_processo_resultado[1] if fase_processo_resultado else None,
        }
        data.append(proaq_dict)

    # Adicionando os dados calculados
    for proaq in data:
        obj = ProaqDadosGerais.objects.get(id=proaq['id'])
        proaq['get_status_label'] = obj.get_status_label()
        proaq['get_unidade_daf_label'] = obj.get_unidade_daf_label()
        proaq['get_modalidade_aquisicao_label'] = obj.get_modalidade_aquisicao_label()
        proaq['get_denominacao_nome'] = obj.get_denominacao_nome()
        proaq['get_usuario_nome'] = obj.get_usuario_nome()
    
    return JsonResponse({
        'data': data,
        'total_processos': total_processos,
        'has_next': proaqs_paginados.has_next(),
        'has_previous': proaqs_paginados.has_previous(),
        'current_page': page
    })

def proaq_exportar(request):
    print("Exportar Processos Aquisitivos")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        status_proaq = data.get('status_proaq')
        unidade_daf = data.get('unidade_daf')
        modalidade_aquisicao = data.get('modalidade_aquisicao')
        fase_evolucao = data.get('fase_evolucao')        
        denominacao = data.get('denominacao')
        
        filters = {'del_status': False}
        if status_proaq:
            filters['status'] = status_proaq
        if unidade_daf:
            filters['unidade_daf'] = unidade_daf
        if modalidade_aquisicao:
            filters['modalidade_aquisicao'] = modalidade_aquisicao
        if denominacao:
            filters['denominacao_id'] = denominacao

        # Obter a fase mais recente de cada ProaqDadosGerais
        ultima_fase = ProaqEvolucao.objects.filter(
            proaq_id=OuterRef('pk'),
            del_status=False
        ).order_by('-fase_numero')

        # Adicionando o campo da última fase ao queryset
        tab_proaqs = ProaqDadosGerais.objects.annotate(
            ultima_fase=Subquery(ultima_fase.values('fase')[:1])
        ).filter(**filters)

        # Filtrar pela fase de evolução, se necessário
        if fase_evolucao:
            tab_proaqs = tab_proaqs.filter(ultima_fase=fase_evolucao)

        tab_proaqs = tab_proaqs.order_by('denominacao')
        data_exportacao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Cria o workbook e adiciona as abas
        wb = Workbook()
        proaq_dados_gerais = wb.active
        proaq_dados_gerais.title = "ProaqDadosGerais"
        proaq_itens = wb.create_sheet(title="ProaqItens")
        proaq_evolucao = wb.create_sheet(title="ProaqEvolucao")
        proaq_tramitacoes = wb.create_sheet(title="ProaqTramitacao")
        
        # Escreve os cabeçalhos
        proaq_dados_gerais.append([
            'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Unidade DAF', 'Modalidade Aquisicao', 'Denominacao', 'Numero Processo SEI', 'Numero ETP',
            'Status', 'Responsavel Tecnico', 'Observacoes Gerais', 'Data Exportação'
        ])
        proaq_itens.append([
            'ID Item', 'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Item', 'Tipo de Cota', 'Produto', 'ID Produto',
            'CMM - Início', 'CMM - Fim', 'CMM - Referência', 'Qtd a ser adquirida', 
            'Cobertura - dias', 'Cobertura - meses',
            'Valor Unitário Estimado - R$', 'Valor Total R$', 'Observações'
            'Data Exportação'
        ])
        proaq_evolucao.append([
            'ID Evolução', 'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'N Fase', 'Fase', 'Data Início', 'Data Fim', 'Dias', 'Observações', 'Data Exportação'
        ])
        proaq_tramitacoes.append([
            'ID Tramitação', 'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Documento SEI', 'Setor', 'Etapa Processo', 
            'Data Entrada', 'Previsao Saida', 'Data Saida', 'Dias',
            'Observacoes', 'Data Exportação'
        ])
        
        for proaq in tab_proaqs:
            registro_data = proaq.registro_data.replace(tzinfo=None)
            ult_atual_data = proaq.ult_atual_data.replace(tzinfo=None)
            # Escreve os dados em ProaqDadosGerais
            proaq_dados_gerais.append([
                proaq.id,
                registro_data, str(proaq.usuario_registro.primeiro_ultimo_nome()),
                ult_atual_data, str(proaq.usuario_atualizacao.primeiro_ultimo_nome()), proaq.log_n_edicoes,
                
                proaq.get_unidade_daf_label(), proaq.get_modalidade_aquisicao_label(), proaq.get_denominacao_nome(),
                proaq.numero_processo_sei, proaq.numero_etp, proaq.get_status_label(),
                proaq.responsavel_tecnico_proaq(),  proaq.observacoes_gerais, data_exportacao
            ])
            
            # Escreve os registros relacionados em ProaqProdutos
            for item in proaq.proaq_item.filter(del_status=False):
                registro_data = item.registro_data.replace(tzinfo=None)
                ult_atual_data = item.ult_atual_data.replace(tzinfo=None)
                proaq_itens.append([
                    item.id, item.proaq.id,
                    registro_data, str(item.usuario_registro.primeiro_ultimo_nome()),
                    ult_atual_data, str(item.usuario_atualizacao.primeiro_ultimo_nome()), item.log_n_edicoes,
                    item.numero_item, item.tipo_cota, item.produto.produto, item.produto.id,
                    item.cmm_data_inicio, item.cmm_data_fim, item.cmm_estimado, item.qtd_a_ser_contratada,
                    item.cobertura_dias(), item.cobertura_meses(),
                    item.valor_unitario_estimado, item.valor_total(),
                    item.observacoes_gerais, data_exportacao
                ])
            
            # Escreve os registros relacionados em ProaqEvolucao
            for evolucao in proaq.proaq_evolucao.filter(del_status=False):
                registro_data = evolucao.registro_data.replace(tzinfo=None)
                ult_atual_data = evolucao.ult_atual_data.replace(tzinfo=None)
                proaq_evolucao.append([
                    evolucao.id, evolucao.proaq.id,
                    registro_data, str(evolucao.usuario_registro.primeiro_ultimo_nome()),
                    ult_atual_data, str(evolucao.usuario_atualizacao.primeiro_ultimo_nome()), evolucao.log_n_edicoes, 
                    evolucao.fase_numero, evolucao.fase_numero,
                    evolucao.data_entrada, evolucao.data_saida, evolucao.total_dias(),
                    evolucao.observacoes_gerais, data_exportacao
                ])
            
            # Escreve os registros relacionados em ProaqTramitacao
            for tramitacao in proaq.proaq_tramitacao.filter(del_status=False):
                registro_data = proaq.registro_data.replace(tzinfo=None)
                ult_atual_data = proaq.ult_atual_data.replace(tzinfo=None)
                proaq_tramitacoes.append([
                    tramitacao.id, tramitacao.proaq.id, 
                    registro_data, str(tramitacao.usuario_registro.primeiro_ultimo_nome()),
                    ult_atual_data, str(tramitacao.usuario_atualizacao.primeiro_ultimo_nome()), tramitacao.log_n_edicoes,
                    tramitacao.documento_sei, tramitacao.setor.setor, tramitacao.etapa_tramitacao(),
                    tramitacao.data_entrada, tramitacao.previsao_saida, tramitacao.data_saida, tramitacao.dias_tramitacao(),
                    tramitacao.observacoes, data_exportacao
                ])
        
        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Processo Aquisitivo",
            model='ProaqDadosGerais',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de processos aquisitivos.",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou dados de Processos Aquisitivos em {current_date_str}."
        )
        log_entry.save()

        # Salva o workbook em um arquivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=proaq.xlsx'
        wb.save(response)
        
        return response
    



#PROAQ DADOS GERAIS
def proaq_ficha(request, proaq_id=None):
    if proaq_id:
        proaq = ProaqDadosGerais.objects.get(id=proaq_id)
    else:
        proaq = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if proaq:
            proaq_form = ProaqDadosGeraisForm(request.POST, instance=proaq)
            novo_proaq = False
        else:
            proaq_form = ProaqDadosGeraisForm(request.POST)
            novo_proaq = True
        
        #Verificar se houve alteração no formulário
        if not proaq_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Passar o objeto Denominação Genérica
        denominacao_id = request.POST.get('denominacao')
        denominacao_instance = DenominacoesGenericas.objects.get(id=denominacao_id)

        #Responsável técnico
        responsavel_id = request.POST.get('responsavel_tecnico')
        if responsavel_id:
            responsavel_instance = Usuario.objects.get(id=responsavel_id)
            modificacoes_post['responsavel_tecnico'] = responsavel_instance       

        #Atualizar os valores no mutable_post
        modificacoes_post['denominacao'] = denominacao_instance

        #Criar o formulário com os dados atualizados
        proaq_form = ProaqDadosGeraisForm(modificacoes_post, instance=proaq_form.instance)
        
        #salvar
        if proaq_form.is_valid():
            #Salvar o produto
            proaq = proaq_form.save(commit=False)
            proaq.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_atualizacao_usuario = proaq.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = proaq.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = proaq.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="ProcessosAquisitivos_DadosGerais",
                model='ProaqDadosGerais',
                model_id=proaq.id,
                item_id=0,
                item_descricao="Salvar edição de Processo Aquisitivo.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Processo Aquisitivo (ID {proaq.id}, Número do Processo SEI: {proaq.numero_processo_sei}, Denominação: {proaq.denominacao.denominacao}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': novo_proaq,
                    'redirect_url': reverse('proaq_ficha', args=[proaq.id]),
                })
        else:
            print("Erro formulário Contrato")
            print(proaq_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

    lista_produtos = []
    tab_proaq_itens = None
    tab_proaq_evolucao = None
    tab_proaq_tramitacoes = None
    if proaq:
        form = ProaqDadosGeraisForm(instance=proaq)

        denominacao_id = proaq.denominacao.id
        tab_produtos = ProdutosFarmaceuticos.objects.filter(del_status=False, denominacao_id=denominacao_id)
        for produto in tab_produtos:
            lista_produtos.append((produto.id, produto.produto))

        tab_proaq_itens = ProaqItens.objects.filter(del_status=False, proaq_id=proaq.id)
        tab_proaq_evolucao = ProaqEvolucao.objects.filter(del_status=False, proaq_id=proaq.id)
        tab_proaq_tramitacoes = ProaqTramitacao.objects.filter(del_status=False, proaq_id=proaq.id)
    else:
        form = ProaqDadosGeraisForm()   

    form_item_proaq = ProaqItensForm()
    form_evolucao_proaq = ProaqEvolucaoForm()
    form_tramitacao_proaq = ProaqTramitacaoForm()

    fases = [
        {'nome': 'Docs Iniciais', 'numero': 1},
        {'nome': 'Pré-Contratual', 'numero': 2},
        {'nome': 'Contratação', 'numero': 3},
        {'nome': 'Execução do Contrato', 'numero': 4},
        {'nome': 'Prestação de Contas', 'numero': 5},
        {'nome': 'Encerrado', 'numero': 6},
    ]
    proaq_fases = []
    
    
    for fase in fases:
        if tab_proaq_evolucao:
            evolucao = tab_proaq_evolucao.filter(fase_numero=fase['numero']).first()
        else:
            evolucao = None
        if evolucao:
            proaq_fases.append([
                fase['nome'], 
                evolucao.data_entrada, 
                evolucao.data_saida,
                evolucao.total_dias(),
                evolucao.id,
            ])
        else:
            proaq_fases.append([fase['nome'], None, None, None, None])

    return render(request, 'processos_aquisitivos/proaq_ficha_dados_gerais.html', {
        'form': form,
        'proaq': proaq,
        'form_item_proaq': form_item_proaq,
        'form_evolucao_proaq': form_evolucao_proaq,
        'form_tramitacao_proaq': form_tramitacao_proaq,
        'proaq_fases': proaq_fases,
        'STATUS_PROAQ': STATUS_PROAQ,
        'lista_fases_evolucao_proaq': FASES_EVOLUCAO_PROAQ,
        'lista_produtos': lista_produtos,
        'tab_proaq_itens': tab_proaq_itens,
        'tab_proaq_evolucao': tab_proaq_evolucao,
        'tab_proaq_tramitacoes': tab_proaq_tramitacoes,
    })

def proaq_dados_gerais_deletar(request, proaq_id=None):
    try:
        proaq = ProaqDadosGerais.objects.get(id=proaq_id)
        proaq.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Processo Aquisitivo deletado com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Processo Aquisitivo_Dados Gerais",
            model='ProaqDadosGerais',
            model_id=proaq.id,
            item_id=0,
            item_descricao="Deleção de processo aquisitivo.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o processo aquisitivo (ID: {proaq.id}) da unidade daf {proaq.unidade_daf} e denominação genérica {proaq.denominacao.denominacao} em {current_date_str}."
        )
        log_entry.save()
        
        return JsonResponse({
            "message": "Processo Aquisitivo deletado com sucesso!",
            'redirect_url': '/proaq/',
            })
    except ProaqDadosGerais.DoesNotExist:
        messages.error(request, "Processo Aquisitivo não encontrado.")
        return JsonResponse({"message": "Processo Aquisitivo não encontrado."})

def proaq_relatorio_pdf(request, proaq_id=None):
    proaq = ProaqDadosGerais.objects.get(id=proaq_id)
    tab_proaq_itens = ProaqItens.objects.filter(del_status=False, proaq_id=proaq.id)
    tab_proaq_tramitacoes = ProaqTramitacao.objects.filter(del_status=False, proaq_id=proaq.id)

    #Evolução do Processo Aquisitivo
    tab_proaq_evolucao = ProaqEvolucao.objects.filter(del_status=False, proaq_id=proaq.id)
    fases = [
        {'nome': 'Docs Iniciais', 'numero': 1},
        {'nome': 'Pré-Contratual', 'numero': 2},
        {'nome': 'Contratação', 'numero': 3},
        {'nome': 'Execução do Contrato', 'numero': 4},
        {'nome': 'Prestação de Contas', 'numero': 5},
        {'nome': 'Encerrado', 'numero': 6},
    ]
    proaq_fases = []
    for fase in fases:
        if tab_proaq_evolucao:
            evolucao = tab_proaq_evolucao.filter(fase_numero=fase['numero']).first()
        else:
            evolucao = None
        if evolucao:
            proaq_fases.append([
                fase['nome'], 
                evolucao.data_entrada, 
                evolucao.data_saida,
                evolucao.total_dias(),
                evolucao.id,
            ])
        else:
            proaq_fases.append([fase['nome'], None, None, None, None])

    #Log Relatório
    usuario_nome = request.user.usuario_relacionado.primeiro_ultimo_nome
    data_hora_atual = datetime.now()
    data_hora = data_hora_atual.strftime('%d/%m/%Y %H:%M:%S')
    
    conteudo = {
        'proaq': proaq,
        'tab_proaq_itens': tab_proaq_itens,
        'tab_proaq_tramitacoes': tab_proaq_tramitacoes,
        'proaq_fases': proaq_fases,
        'usuario': usuario_nome,
        'data_hora': data_hora,
    }
    return render(request, 'processos_aquisitivos/proaq_relatorio.html', conteudo)




#PROAQ ITENS
def proaq_item_salvar(request, proaq_item_id=None):
    if proaq_item_id:
        proaq_item = ProaqItens.objects.get(id=proaq_item_id)
    else:
        proaq_item = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if proaq_item:
            proaq_item_form = ProaqItensForm(request.POST, instance=proaq_item)
            novo_proaq_item = False
        else:
            proaq_item_form = ProaqItensForm(request.POST)
            novo_proaq_item = True
        
        #Verificar se houve alteração no formulário
        if not proaq_item_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Objeto
        produto_id = request.POST.get('produto')
        produto_instance = ProdutosFarmaceuticos.objects.get(id=produto_id)
        
        #Processo Aquisitivo
        proaq_id = request.POST.get('id_proaq_item_hidden')
        proaq_instance = ProaqDadosGerais.objects.get(id=proaq_id)

        #Valor Unitário
        valor_unitario = request.POST.get('valor_unitario_estimado')
        valor_unitario = valor_unitario.replace('R$', '').replace('.', '').replace(',', '.').strip()
        valor_unitario = float(valor_unitario)

        #CMM de Referência
        cmm_referencia = request.POST.get('cmm_estimado')
        cmm_referencia = float(cmm_referencia.replace('.', ''))

        #Qtd a ser Contratada
        qtd_a_ser_contratada = request.POST.get('qtd_a_ser_contratada')
        qtd_a_ser_contratada = float(qtd_a_ser_contratada.replace('.', ''))

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['produto'] = produto_instance
        modificacoes_post['proaq'] = proaq_instance
        modificacoes_post['valor_unitario_estimado'] = valor_unitario
        modificacoes_post['qtd_a_ser_contratada'] = qtd_a_ser_contratada
        modificacoes_post['cmm_estimado'] = cmm_referencia

        #Criar o formulário com os dados atualizados
        proaq_item_form = ProaqItensForm(modificacoes_post, instance=proaq_item_form.instance)

        #salvar
        if proaq_item_form.is_valid():
            #Salvar o produto
            proaq_item = proaq_item_form.save(commit=False)
            proaq_item.save(current_user=request.user.usuario_relacionado)

            #id do processo aquisitivo e do item
            proaq_id = proaq_item.proaq.id

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="ProcessosAquisitivos_Itens",
                model='ProaqItens',
                model_id=proaq_item.id,
                item_id=0,
                item_descricao="Salvar edição de Item do Processo Aquisitivo.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Item do Processo Aquisitivo (ID {proaq_item.id}, Processo Aquisitivo: {proaq_item.proaq.numero_processo_sei}, Produto: {proaq_item.produto.produto}) em {current_date_str}."
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': novo_proaq_item,
                    'redirect_url': reverse('proaq_ficha', args=[proaq_id]),
                    'id_proaq_item': proaq_item.id,
                })
        else:
            print("Erro formulário Parcela do Contrato")
            print(proaq_item_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def proaq_item_modal(request, proaq_item_id=None):
    try:
        item = ProaqItens.objects.get(id=proaq_item_id)
        produto = item.produto.id
        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'tipo_cota': item.tipo_cota,
            'numero_item': item.numero_item,
            'produto': produto,
            'data_inicio': item.cmm_data_inicio,
            'data_fim': item.cmm_data_fim,
            'cmm_estimado': item.cmm_estimado,
            'qtd_a_ser_contratada': item.qtd_a_ser_contratada,
            'valor_unitario_estimado': item.valor_unitario_estimado,
            'observacoes': item.observacoes_gerais,
        }
        return JsonResponse(data)
    except ProaqItens.DoesNotExist:
        return JsonResponse({'error': 'Item do Processo Aquisitivo não encontrado.'}, status=404)

def proaq_item_deletar(request, proaq_item_id=None):
    try:
        proaq_item = ProaqItens.objects.get(id=proaq_item_id)
        proaq_item.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="ProcessosAquisitivos_Itens",
            model='ProaqItens',
            model_id=proaq_item.id,
            item_id=0,
            item_descricao="Deleção da Parcela de Contrato.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Item do Processo Aquisitivo (ID {proaq_item.id}, Processo SEI: {proaq_item.proaq.numero_processo_sei}, Produto: {proaq_item.produto.produto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Item do Processo deletado com sucesso!"
            })
    except ProaqItens.DoesNotExist:
        return JsonResponse({
            "message": "Item do Processo não encontrado."
            })



#PROAQ EVOLUÇÃO
def proaq_ficha_evolucao(request, proaq_evolucao_id=None):
    if proaq_evolucao_id:
        proaq_evolucao = ProaqEvolucao.objects.get(id=proaq_evolucao_id)
    else:
        proaq_evolucao = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if proaq_evolucao:
            proaq_evolucao_form = ProaqEvolucaoForm(request.POST, instance=proaq_evolucao)
            novo_proaq_evolucao = False
        else:
            proaq_evolucao_form = ProaqEvolucaoForm(request.POST)
            novo_proaq_evolucao = True
        
        #Verificar se houve alteração no formulário
        if not proaq_evolucao_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })
        
        #Processo Aquisitivo
        proaq_id = request.POST.get('id_evolucaoproaq_proaq_hidden')
        proaq_instance = ProaqDadosGerais.objects.get(id=proaq_id)

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['proaq'] = proaq_instance

        #Criar o formulário com os dados atualizados
        proaq_evolucao_form = ProaqEvolucaoForm(modificacoes_post, instance=proaq_evolucao_form.instance)

        #salvar
        if proaq_evolucao_form.is_valid():
            #Salvar o produto
            proaq_evolucao = proaq_evolucao_form.save(commit=False)
            proaq_evolucao.save(current_user=request.user.usuario_relacionado)

            #id do processo aquisitivo
            proaq_id = proaq_evolucao.proaq.id

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="ProcessosAquisitivos_Evolucao",
                model='ProaqEvolucao',
                model_id=proaq_evolucao.id,
                item_id=0,
                item_descricao="Salvar edição da Evolução do Processo Aquisitivo.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou a Evolução do Processo Aquisitivo (ID {proaq_evolucao.id}, Processo Aquisitivo: {proaq_evolucao.proaq.numero_processo_sei} em {current_date_str}."
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': novo_proaq_evolucao,
                    'redirect_url': reverse('proaq_ficha', args=[proaq_id]),
                    'id_evolucao_proaq': proaq_evolucao.id,
                })
        else:
            print("Erro formulário Parcela do Contrato")
            print(proaq_evolucao_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def proaq_evolucao_modal(request, proaq_evolucao_id=None):
    try:
        item = ProaqEvolucao.objects.get(id=proaq_evolucao_id)
        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'fase_numero': item.fase_numero,
            'fase': item.fase,
            'data_entrada': item.data_entrada,
            'data_saida': item.data_saida,
            'observacoes': item.observacoes_gerais,
        }
        print('Data da Entrada: ', item.data_entrada)
        print('Data da Saída: ', item.data_saida)
        return JsonResponse(data)
    except ProaqEvolucao.DoesNotExist:
        return JsonResponse({'error': 'Evolução do Processo Aquisitivo não encontrado.'}, status=404)

def proaq_evolucao_deletar(request, proaq_evolucao_id=None):
    try:
        proaq_evolucao = ProaqEvolucao.objects.get(id=proaq_evolucao_id)
        proaq_evolucao.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="ProcessosAquisitivos_Evolucao",
            model='ProaqEvolucao',
            model_id=proaq_evolucao.id,
            item_id=0,
            item_descricao="Deleção da Evolução do Processo Aquisitivo.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a Evolução do Processo Aquisitivo (ID {proaq_evolucao.id}, Processo SEI: {proaq_evolucao.proaq.numero_processo_sei}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Evolução do Processo Aquisitivo deletado com sucesso!"
            })
    except ProaqEvolucao.DoesNotExist:
        return JsonResponse({
            "message": "Evolução do Processo Aquisitivo não encontrado."
            })



#PROAQ TRAMITACAO
def proaq_tramitacao_salvar(request, tramitacao_id=None):
    if tramitacao_id:
        proaq_tramitacao = ProaqTramitacao.objects.get(id=tramitacao_id)
    else:
        proaq_tramitacao = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if proaq_tramitacao:
            proaq_tramitacao_form = ProaqTramitacaoForm(request.POST, instance=proaq_tramitacao)
            novo_proaq_tramitacao = False
        else:
            proaq_tramitacao_form = ProaqTramitacaoForm(request.POST)
            novo_proaq_tramitacao = True
        
        #Verificar se houve alteração no formulário
        if not proaq_tramitacao_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })
        
        #Processo Aquisitivo
        proaq_id = request.POST.get('proaq_hidden')
        proaq_instance = ProaqDadosGerais.objects.get(id=proaq_id)
        
        #Setor
        setor_id = request.POST.get('setor')
        if setor_id:
            setor_instance = PROAQ_AREA_MS.objects.get(id=setor_id)
        else:
            setor_instance = None

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['proaq'] = proaq_instance
        modificacoes_post['setor'] = setor_instance

        #Criar o formulário com os dados atualizados
        proaq_tramitacao_form = ProaqTramitacaoForm(modificacoes_post, instance=proaq_tramitacao_form.instance)

        #salvar
        if proaq_tramitacao_form.is_valid():
            #Salvar o produto
            proaq_tramitacao = proaq_tramitacao_form.save(commit=False)
            proaq_tramitacao.save(current_user=request.user.usuario_relacionado)

            #id do processo aquisitivo
            proaq_id = proaq_tramitacao.proaq.id

            if proaq_tramitacao.etapa_processo:
                etapa_processo = proaq_tramitacao.etapa_processo
            else:
                etapa_processo = proaq_tramitacao.etapa_processo_outro

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="ProcessosAquisitivos_Tramitacao",
                model='ProaqTramitacao',
                model_id=proaq_tramitacao.id,
                item_id=0,
                item_descricao="Salvar edição da Tramitação do Processo Aquisitivo.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou a Tramitacao do Processo Aquisitivo (ID {proaq_tramitacao.id}, Etapa: {etapa_processo}, Setor: {proaq_tramitacao.setor} Processo Aquisitivo: {proaq_tramitacao.proaq.numero_processo_sei} em {current_date_str}."
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'novo': novo_proaq_tramitacao,
                    'redirect_url': reverse('proaq_ficha', args=[proaq_id]),
                    'id_tramitacao_proaq': proaq_tramitacao.id,
                })
        else:
            print("Erro formulário Tramitação do Processo Aquisitivo")
            print(proaq_tramitacao_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

def proaq_tramitacao_modal(request, tramitacao_id=None):
    try:
        item = ProaqTramitacao.objects.get(id=tramitacao_id)
        if item.etapa_processo:
            etapa_check_outro = False
        else:
            etapa_check_outro = True
        print('Observações: ', item.observacoes)
        data = {
            #log
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            
            #dados do processo aquisitivo
            'documento_sei': item.documento_sei,
            'setor': item.setor.id,
            'etapa_processo': item.etapa_processo,
            'etapa_processo_outro': item.etapa_processo_outro,
            'etapa_check_outro': etapa_check_outro,
            
            #datas
            'data_entrada': item.data_entrada,
            'data_previsao': item.previsao_saida,
            'data_saida': item.data_saida,

            #observacoes
            'observacoes': item.observacoes,
        }
        return JsonResponse(data)
    except ProaqTramitacao.DoesNotExist:
        return JsonResponse({'error': 'Tramitação do Processo Aquisitivo não encontrada.'}, status=404)

def proaq_tramitacao_deletar(request, tramitacao_id=None):
    try:
        tramitacao = ProaqTramitacao.objects.get(id=tramitacao_id)
        tramitacao.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="ProcessosAquisitivos_Tramitacao",
            model='ProaqTramitacao',
            model_id=tramitacao.id,
            item_id=0,
            item_descricao="Deleção da Tramitação do Processo Aquisitivo.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a Tramitação do Processo Aquisitivo (ID {tramitacao.id}, Tramitação: {tramitacao.etapa_tramitacao()} Processo SEI: {tramitacao.proaq.numero_processo_sei}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Tramitação do Processo Aquisitivo deletado com sucesso!"
            })
    except ProaqEvolucao.DoesNotExist:
        return JsonResponse({
            "message": "Tramitação do Processo Aquisitivo não encontrado."
            })
    