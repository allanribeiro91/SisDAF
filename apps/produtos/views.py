from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage
from django.contrib.auth.decorators import login_required
from django.contrib import auth, messages
from django.contrib.auth.models import User
from apps.usuarios.models import Usuario
from apps.produtos.models import DenominacoesGenericas, ProdutosFarmaceuticos, Tags, ListaATC, ProdutosTags
from apps.produtos.forms import DenominacoesGenericasForm, ProdutosFarmaceuticosForm
from setup.choices import TIPO_PRODUTO, FORMA_FARMACEUTICA, STATUS_INCORPORACAO, CONCENTRACAO_TIPO, YES_NO, CLASSIFICACAO_AWARE
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import json

@login_required
def produtos(request):
    produtos = ProdutosFarmaceuticos.objects.filter(del_status=False)
    numero_produtos = produtos.count()
    conteudo = {
        'produtos': produtos,
        'TIPO_PRODUTO': TIPO_PRODUTO,
        'numero_produtos': numero_produtos,
    }
    return render(request, 'produtos/produtos.html', conteudo)


@login_required
def produtos_ficha(request, product_id=None):

    tags_selecionadas = []
    if product_id:
         try:
             produto = ProdutosFarmaceuticos.objects.get(id=product_id)
             tags_selecionadas = list(produto.produto_tag.active().values('tag_id', 'tag'))
         except ProdutosFarmaceuticos.DoesNotExist:
             messages.error(request, "Produto não encontrado.")
             return redirect('produtos')
    else:
         produto = None  # Preparando para criar um novo produto


    if request.method == 'POST':
        if produto:
            produto_form = ProdutosFarmaceuticosForm(request.POST, instance=produto)
        else:
            produto_form = ProdutosFarmaceuticosForm(request.POST)

        #Conferir se produto e tipo de produto foram preenchidos
        nome_produto = request.POST.get('produto')
        tipo_produto = request.POST.get('tipo_produto')
        if not nome_produto or not tipo_produto:
            messages.error(request, "O nome do produto é obrigatório!")
            if produto:
                return redirect('produtos_ficha', product_id=produto.id)
            else:
                return redirect('novo_produto')

        #Verificar se houve alteração no formulário
        if not produto_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if produto:
                return redirect('produtos_ficha', product_id=produto.id)
            else:
                return redirect('novo_produto')

        if produto_form.is_valid():
            #Verificar se já existe a denominação na base [
            nome_produto = produto_form.cleaned_data.get('produto')
            produto_existente = ProdutosFarmaceuticos.objects.filter(produto=nome_produto)
            
            #Se estivermos atualizando uma denominação existente, excluímos essa denominação da verificação
            if produto:
                produto_existente = produto_existente.exclude(id=produto.id)

            if produto_existente.exists():
                messages.error(request, "Já existe um produto com esse nome. Não foi possível salvar.")
                if produto:
                    return redirect('produtos_ficha', product_id=produto.id)
                else:
                    return redirect('novo_produto')

            #Salvar o produto
            produto = produto_form.save(commit=False)
            produto.save(current_user=request.user.usuario_relacionado)
            messages.success(request, f"Dados atualizados com sucesso!")
            
            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'registro_data': produto.registro_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_registro': produto.usuario_registro.dp_nome_completo,
                    'ult_atual_data': produto.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_atualizacao': produto.usuario_atualizacao.dp_nome_completo,
                    'log_n_edicoes': produto.log_n_edicoes,
                })
   
        else:
            messages.error(request, "Formulário inválido")
            print("Erro formulário produto")
            print(produto_form.errors)

    denominacoes_genericas = DenominacoesGenericas.objects.values_list('id', 'denominacao', 'tipo_produto')
    lista_atc = list(ListaATC.objects.values_list('codigo', 'descricao'))
    lista_atc.insert(0, ('Não informado', 'Não informado'))    
    tags_produtos = json.dumps(list(Tags.objects.values_list('id','tag')))
    form = ProdutosFarmaceuticosForm(instance=produto)
    
    return render(request, 'produtos/produtos_ficha.html', {
        'product_id': product_id,
        'produto': produto,
        'form': form,
        'denominacoes_genericas': denominacoes_genericas,
        'lista_atc': lista_atc,
        'tags_produtos': tags_produtos,
        'tags_selecionadas': tags_selecionadas,
        'TIPO_PRODUTO': TIPO_PRODUTO,
        'FORMA_FARMACEUTICA': FORMA_FARMACEUTICA, 
        'STATUS_INCORPORACAO': STATUS_INCORPORACAO,
        'CONCENTRACAO_TIPO': CONCENTRACAO_TIPO,
        'YES_NO': YES_NO,
        'CLASSIFICACAO_AWARE': CLASSIFICACAO_AWARE,
    })

@login_required
def delete_produto(request, product_id):
    try:
        produto = ProdutosFarmaceuticos.objects.get(id=product_id)
        produto.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Produto deletado com sucesso.")
        return JsonResponse({"message": "Produto deletado com sucesso!"})
    except ProdutosFarmaceuticos.DoesNotExist:
        messages.error(request, "Produto não encontrado.")    
    return redirect('produtos')

@login_required
def get_filtros_produtos(request):
    tipo_produto = request.GET.get('tipo_produto', None)
    produto = request.GET.get('produto', None)
    basico = request.GET.get('basico', None)
    especializado = request.GET.get('especializado', None)
    estrategico = request.GET.get('estrategico', None)
    farmacia_popular = request.GET.get('farmacia_popular', None)
    hospitalar = request.GET.get('hospitalar', None)

    filters = {}
    filters['del_status'] = False
    if tipo_produto:
        filters['denominacao__tipo_produto'] = tipo_produto
    if produto:
        filters['produto__icontains'] = produto
    if basico:
        filters['comp_basico'] = 1
    if especializado:
        filters['comp_especializado'] = 1
    if estrategico:
        filters['comp_estrategico'] = 1
    if farmacia_popular:
        filters['disp_farmacia_popular'] = 1
    if hospitalar:
        filters['hospitalar'] = 1
    
    produtos = ProdutosFarmaceuticos.objects.filter(**filters)
    numero_produtos = produtos.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(produtos, 100)  # Mostra 100 denominações por página
    try:
        produtos_paginados = paginator.page(page)
    except EmptyPage:
        produtos_paginados = paginator.page(paginator.num_pages)

    #data = list(produtos_paginados.object_list.values())
    data = list(produtos_paginados.object_list.values('id', 'denominacao__tipo_produto', 'produto', 'comp_basico', 'comp_especializado', 'comp_estrategico', 'disp_farmacia_popular', 'hospitalar'))

    return JsonResponse({
        'data': data,
        'numero_produtos': numero_produtos,
        'has_next': produtos_paginados.has_next(),
        'has_previous': produtos_paginados.has_previous(),
        'current_page': page
    })

@login_required
def exportar_produtos(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        tipo_produto = data.get('tipo_produto')
        produto = data.get('produto')
        basico = data.get('basico')
        especializado = data.get('especializado')
        estrategico = data.get('estrategico')
        farmacia_popular = data.get('farmacia_popular')
        hospitalar = data.get('hospitalar')
    
        filters = {}
        filters['del_status'] = False
        if tipo_produto:
            filters['denominacao__tipo_produto'] = tipo_produto
        if produto:
            filters['produto__icontains'] = produto
        if basico:
            filters['unidade_basico'] = basico
        if especializado:
            filters['unidade_especializado'] = especializado
        if estrategico:
            filters['unidade_estrategico'] = estrategico
        if farmacia_popular:
            filters['unidade_farm_popular'] = farmacia_popular
        if hospitalar:
            filters['hospitalar'] = hospitalar
        
        produtos = ProdutosFarmaceuticos.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "denominacoes_genericas"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização',
        'N Edições', 'Tipo Produto', 'Produto Farmacêutico', 'Concentração', 'Forma Farmacêutica', 'Oncológico', 'Biológico',
        'AWaRe', 'ATC - Código', 'ATC - Descrição', 'Incorporação SUS', 'Data Incorporação', 'Portaria Incorp.',
        'Link Portaria Incorp.', 'Data Exclusão', 'Portaria Exclusão', 'Link Exclusão',
        'Comp. Básico', 'Comp. Especializado', 'Comp. Estratégico', 'Disp. Farmácia Popular', 'Hospitalar',
        'SIGTAP', 'SIGTAP Cód.', 'SIGTAP Nome', 'SISMAT', 'SISMAT Cód.', 'SISMAT Nome', 'CATMAT', 'CATMAT Cód.', 'CATMAT Nome',
        'OBM', 'OBM Cód.', 'OBM Nome', 'Observações gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 15

        # Adicionar dados da tabela
        for row_num, produto in enumerate(produtos, 2):
            ws.cell(row=row_num, column=1, value=produto.id)
            ws.cell(row=row_num, column=2, value=str(produto.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(produto.usuario_atualizacao.primeiro_ultimo_nome()))
            registro_data = produto.registro_data.replace(tzinfo=None)
            ult_atual_data = produto.ult_atual_data.replace(tzinfo=None)
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=produto.log_n_edicoes)
            ws.cell(row=row_num, column=7, value=produto.denominacao.tipo_produto)
            ws.cell(row=row_num, column=8, value=produto.produto)
            ws.cell(row=row_num, column=9, value=produto.concentracao)
            ws.cell(row=row_num, column=10, value=produto.get_forma_farmaceutica_display())
            ws.cell(row=row_num, column=11, value=produto.oncologico)
            ws.cell(row=row_num, column=12, value=produto.biologico)
            ws.cell(row=row_num, column=13, value=produto.aware)
            ws.cell(row=row_num, column=14, value=produto.atc)
            ws.cell(row=row_num, column=15, value=produto.atc_descricao)
            ws.cell(row=row_num, column=16, value=produto.incorp_status)
            ws.cell(row=row_num, column=17, value=produto.incorp_portaria)
            ws.cell(row=row_num, column=18, value=produto.incorp_link)
            ws.cell(row=row_num, column=19, value=produto.exclusao_data)
            ws.cell(row=row_num, column=20, value=produto.exclusao_portaria)
            ws.cell(row=row_num, column=21, value=produto.exclusao_link)
            ws.cell(row=row_num, column=22, value=produto.comp_basico)
            ws.cell(row=row_num, column=23, value=produto.comp_especializado)
            ws.cell(row=row_num, column=24, value=produto.comp_estrategico)
            ws.cell(row=row_num, column=25, value=produto.disp_farmacia_popular)
            ws.cell(row=row_num, column=26, value=produto.hospitalar)
            ws.cell(row=row_num, column=27, value=produto.sigtap_possui)
            ws.cell(row=row_num, column=28, value=produto.sigtap_codigo)
            ws.cell(row=row_num, column=29, value=produto.sigtap_nome)
            ws.cell(row=row_num, column=30, value=produto.sismat_possui)
            ws.cell(row=row_num, column=31, value=produto.sismat_codigo)
            ws.cell(row=row_num, column=32, value=produto.sismat_nome)
            ws.cell(row=row_num, column=33, value=produto.catmat_possui)
            ws.cell(row=row_num, column=34, value=produto.catmat_codigo)
            ws.cell(row=row_num, column=35, value=produto.catmat_nome)
            ws.cell(row=row_num, column=36, value=produto.obm_possui)
            ws.cell(row=row_num, column=37, value=produto.obm_codigo)
            ws.cell(row=row_num, column=38, value=produto.obm_nome)
            ws.cell(row=row_num, column=39, value=produto.observacoes_gerais)
            ws.cell(row=row_num, column=15, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_denominacoes.xlsx"'
        response.write(output.getvalue())
        return response

@login_required
def salvar_tags(request, product_id):
    if request.method == "POST":
        produto = get_object_or_404(ProdutosFarmaceuticos, pk=product_id)

        # Decodificar as tags enviadas de uma string JSON para uma lista de dicionários
        tags_selecionadas_json = request.POST.get('tags')
        tags_selecionadas_list = json.loads(tags_selecionadas_json)

        # Obter apenas os IDs das tags selecionadas para comparação
        tags_selecionadas_ids = [tag["id"] for tag in tags_selecionadas_list]
        
        # Tags atuais do produto
        tags_banco_ids = list(produto.produto_tag.active().values_list('tag_id', flat=True))
        
        # Tags para salvar
        tags_salvar = [tag for tag in tags_selecionadas_list if tag["id"] not in tags_banco_ids]

        # Usuario
        usuario_instance = request.user.usuario_relacionado
        
        # Tags para adicionar
        for tag_data in tags_salvar:
            tag_instance = ProdutosTags(
                usuario_registro=usuario_instance,
                usuario_atualizacao=usuario_instance,
                produto=produto,
                tag=tag_data["value"],
                tag_id=tag_data["id"]
            )
            tag_instance.save()

        # Tags para deletar (soft delete)
        tags_deletar = [tag_id for tag_id in tags_banco_ids if tag_id not in tags_selecionadas_ids]
        
        print("TAGS ATUAIS: ", tags_banco_ids)
        print("TAGS SELECIONADAS: ", tags_selecionadas_ids)
        print("TAGS PARA DELETAR: ", tags_deletar)
        for tag_id in tags_deletar:
            tag_instance = ProdutosTags.objects.filter(produto=produto, tag_id=tag_id, del_status=False).first()
            print("Chamando soft_delete para tag_id:", tag_id)
            if tag_instance:
                tag_instance.soft_delete(usuario_instance)

        return JsonResponse({"status": "success"})
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"})





@login_required
def denominacoes(request):
    denominacoes = DenominacoesGenericas.objects.filter(del_status=False)
    numero_denominacoes = denominacoes.count()
    context = {
        'denominacoes': denominacoes,
        'TIPO_PRODUTO': TIPO_PRODUTO,
        'numero_denominacoes': numero_denominacoes,
    }
    return render(request, 'produtos/denominacoes.html', context)



@login_required
def denominacoes_ficha(request, denominacao_id=None):

    if denominacao_id:
        try:
            denominacao = DenominacoesGenericas.objects.get(id=denominacao_id)
        except DenominacoesGenericas.DoesNotExist:
            messages.error(request, "Denominação não encontrada.")
            return redirect('denominacoes')
    else:
        denominacao = None  # Preparando para criar uma nova denominação

    if request.method == 'POST':
        if denominacao:
            denominacao_form = DenominacoesGenericasForm(request.POST, instance=denominacao)
        else:
            denominacao_form = DenominacoesGenericasForm(request.POST)

        #Conferir se denominação e tipo de produto foram preenchidos
        nome_denominacao = request.POST.get('denominacao')
        tipo_produto = request.POST.get('tipo_produto')
        if not nome_denominacao or not tipo_produto:
            messages.error(request, "O nome da denominação genérica e o tipo de produto são obrigatórios!")
            if denominacao:
                return redirect('denominacoes_ficha', denominacao_id=denominacao.id)
            else:
                return redirect('nova_denominacao')

        #Verificar se houve alteração no formulário
        if not denominacao_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if denominacao:
                return redirect('denominacoes_ficha', denominacao_id=denominacao.id)
            else:
                return redirect('nova_denominacao')

        if denominacao_form.is_valid():
            #Verificar se já existe a denominação na base [
            nome_denominacao = denominacao_form.cleaned_data.get('denominacao')
            denominacao_existente = DenominacoesGenericas.objects.filter(denominacao=nome_denominacao)
            
            #Se estivermos atualizando uma denominação existente, excluímos essa denominação da verificação
            if denominacao:
                denominacao_existente = denominacao_existente.exclude(id=denominacao.id)

            if denominacao_existente.exists():
                messages.error(request, "Já existe uma denominação com esse nome. Não foi possível salvar.")
                if denominacao:
                    return redirect('denominacoes_ficha', denominacao_id=denominacao.id)
                else:
                    return redirect('nova_denominacao')

            #Salvar a denominação
            denominacao = denominacao_form.save(commit=False)
            denominacao.save(current_user=request.user.usuario_relacionado)
            messages.success(request, f"Dados atualizados com sucesso!")
            
            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'registro_data': denominacao.registro_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_registro': denominacao.usuario_registro.dp_nome_completo,
                    'ult_atual_data': denominacao.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_atualizacao': denominacao.usuario_atualizacao.dp_nome_completo,
                    'log_n_edicoes': denominacao.log_n_edicoes,
                })           
            
        else:
            messages.error(request, "Formulário inválido")
            print("Erro formulário denominação")
            print(denominacao_form.errors)

    form = DenominacoesGenericasForm(instance=denominacao)
    return render(request, 'produtos/denominacoes_ficha.html', {
        'denominacao': denominacao,
        'form': form,
        'TIPO_PRODUTO': TIPO_PRODUTO,
    })


@login_required
def delete_denominacao(request, denominacao_id):
    try:
        denominacao = DenominacoesGenericas.objects.get(id=denominacao_id)
        denominacao.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Denominação deletada com sucesso!")
        return JsonResponse({"message": "Denominação deletada com sucesso!"})
    except DenominacoesGenericas.DoesNotExist:
        messages.error(request, "Denominação não encontrada.")    
    return redirect('denominacoes')


@login_required
def get_filtros_denominacoes(request):
    tipo_produto = request.GET.get('tipo_produto', None)
    denominacao = request.GET.get('denominacao', None)
    basico = request.GET.get('basico', None)
    especializado = request.GET.get('especializado', None)
    estrategico = request.GET.get('estrategico', None)
    farmacia_popular = request.GET.get('farmacia_popular', None)
    hospitalar = request.GET.get('hospitalar', None)

    filters = {}
    filters['del_status'] = False
    if tipo_produto:
        filters['tipo_produto'] = tipo_produto
    if denominacao:
        filters['denominacao__icontains'] = denominacao
    if basico:
        filters['unidade_basico'] = 1
    if especializado:
        filters['unidade_especializado'] = 1
    if estrategico:
        filters['unidade_estrategico'] = 1
    if farmacia_popular:
        filters['unidade_farm_popular'] = 1
    if hospitalar:
        filters['hospitalar'] = 1
    
    denominacoes = DenominacoesGenericas.objects.filter(**filters)
    numero_denominacoes = denominacoes.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(denominacoes, 100)  # Mostra 100 denominações por página
    try:
        denominacoes_paginados = paginator.page(page)
    except EmptyPage:
        denominacoes_paginados = paginator.page(paginator.num_pages)

    data = list(denominacoes_paginados.object_list.values())

    return JsonResponse({
        'data': data,
        'numero_denominacoes': numero_denominacoes,
        'has_next': denominacoes_paginados.has_next(),
        'has_previous': denominacoes_paginados.has_previous(),
        'current_page': page
    })


@login_required
def exportar_denominacoes(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        tipo_produto = data.get('tipo_produto')
        denominacao = data.get('denominacao')
        basico = data.get('basico')
        especializado = data.get('especializado')
        estrategico = data.get('estrategico')
        farmacia_popular = data.get('farmacia_popular')
        hospitalar = data.get('hospitalar')
    
        filters = {}
        filters['del_status'] = False
        if tipo_produto:
            filters['tipo_produto'] = tipo_produto
        if denominacao:
            filters['denominacao__icontains'] = denominacao
        if basico:
            filters['unidade_basico'] = basico
        if especializado:
            filters['unidade_especializado'] = especializado
        if estrategico:
            filters['unidade_estrategico'] = estrategico
        if farmacia_popular:
            filters['unidade_farm_popular'] = farmacia_popular
        if hospitalar:
            filters['hospitalar'] = hospitalar
        

        denominacoes = DenominacoesGenericas.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "denominacoes_genericas"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização',
        'Denominação', 'Tipo de Produto', 'Unidade Básico', 'Unidade Especializado', 'Unidade Estratégico',
        'Unidade Farmácia Popular', 'Hospitalar', 'Observações Gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 15

        # Adicionar dados da tabela
        for row_num, denominacao in enumerate(denominacoes, 2):
            ws.cell(row=row_num, column=1, value=denominacao.id)
            ws.cell(row=row_num, column=2, value=str(denominacao.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(denominacao.usuario_atualizacao.primeiro_ultimo_nome()))
            registro_data = denominacao.registro_data.replace(tzinfo=None)
            ult_atual_data = denominacao.ult_atual_data.replace(tzinfo=None)
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=denominacao.denominacao)
            ws.cell(row=row_num, column=7, value=denominacao.tipo_produto)
            ws.cell(row=row_num, column=8, value=denominacao.unidade_basico)
            ws.cell(row=row_num, column=9, value=denominacao.unidade_especializado)
            ws.cell(row=row_num, column=10, value=denominacao.unidade_estrategico)
            ws.cell(row=row_num, column=11, value=denominacao.unidade_farm_popular)
            ws.cell(row=row_num, column=12, value=denominacao.hospitalar)
            ws.cell(row=row_num, column=13, value=denominacao.observacoes_gerais)
            ws.cell(row=row_num, column=14, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_denominacoes.xlsx"'
        response.write(output.getvalue())
        return response





